from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from tool.common.project_weekly_audit.models import (
    AuditBatch,
    AuditExecutionContext,
    AuditFinding,
    AuditPeriod,
    AuditStatus,
    ProjectAudit,
    ProjectCandidate,
    ProjectCollectionFilter,
    ProductLine,
    UPDATE_MATRIX_POINTS,
)
from tool.common.project_weekly_audit.report import (
    export_project_audit_xlsx,
    export_project_audit_xlsx_by_product_line,
)


TZ = ZoneInfo("Asia/Shanghai")


def test_xlsx_writes_invalid_reason_in_its_cell_without_repeating_ps(tmp_path):
    project = ProjectCandidate(
        "1", "M314", "Muffin314", "https://c/status", "https://c/root",
        2025, "A", "NORMAL", (2025, 2026),
    )
    findings = []
    for index, point in enumerate(UPDATE_MATRIX_POINTS):
        invalid = index in {2, 3, 4}
        missing_region = index == 4
        findings.append(AuditFinding(
            project.project_id,
            point.label.split(".", 1)[0],
            point.rule_id,
            AuditStatus.INVALID_FORMAT if invalid else (
                AuditStatus.UPDATED if index % 2 else AuditStatus.NOT_UPDATED
            ),
            "格式有误" if missing_region else (
                "PermissionError" if invalid else "timestamp"
            ),
            page_url="https://c/page",
            explanation=(
                "格式有误：查询不到Task Arrangement of Important Test（Must give ETA）"
                if missing_region else (
                    "pageId=42; HTTP 403" if invalid else ""
                )
            ),
        ))
    batch = AuditBatch(
        "matrix1",
        AuditPeriod(
            datetime(2026, 7, 27, tzinfo=TZ),
            datetime(2026, 7, 31, tzinfo=TZ),
        ),
        datetime(2026, 7, 31, tzinfo=TZ),
        [ProjectAudit(project, findings, "Alice Bob")],
        ProjectCollectionFilter("DOPL + SDPL", (2025, 2026)),
        AuditExecutionContext("manual"),
    )

    sheet = load_workbook(
        export_project_audit_xlsx(batch, tmp_path / "audit.xlsx"),
    ).active

    assert sheet.title == "Project Weekly Audit"
    assert [cell.value for cell in sheet[1]][:6] == [
        "Support Mode", "A", "Project Status", "NORMAL",
        "审查周期", "2026-07-27 - 2026-07-31",
    ]
    assert [cell.value for cell in sheet[2]] == [
        "Owner", "年份", "项目名", "项目链接",
        *(point.label for point in UPDATE_MATRIX_POINTS), "PS",
    ]
    assert [point.label for point in UPDATE_MATRIX_POINTS] == [
        "Project Status Report.Highlights",
        "Project Status Report.Impact issues",
        "Basic Information.Test Information.Phase Status（当前阶段测试状态）",
        "Basic Information.Test Information.项目整体状态Summary",
        "Basic Information.Test Information.Task Arrangement of Important Test（Must give ETA）",
        "Basic Information.Test Information.Blocking QA Testing Items",
        "Basic Information.Test Information.Test Plan.Category",
        "Basic Information.Test Information.Test Environment Setup and Precautions.测试环境搭建以及注意事项",
        "Basic Information.Test Information.Summary of Experience and Typical Cases",
        "Basic Information.Test Information.Test Report Store",
    ]
    assert sheet.max_row == 3
    assert sheet["A3"].value == "Alice Bob"
    assert sheet["B3"].value == "2025, 2026"
    assert sheet["C3"].value == "Muffin314"
    assert sheet["D3"].value == "https://c/root"
    assert sheet["D3"].hyperlink.target == "https://c/root"
    assert sheet["G3"].value == "格式有误：pageId=42; HTTP 403"
    assert sheet["H3"].value == "格式有误：pageId=42; HTTP 403"
    assert sheet["I3"].value == "格式有误：查询不到Task Arrangement of Important Test（Must give ETA）"
    assert sheet["O3"].value is None
    assert sheet.freeze_panes == "A3"
    assert sheet["A1"].fill.fgColor.rgb == "FFD9EAF7"
    assert sheet["A2"].fill.fgColor.rgb == "FF1F4E78"
    assert sheet["A2"].font.bold is True
    assert sheet["A2"].font.color.rgb == "FFFFFFFF"
    assert sheet["E3"].fill.fgColor.rgb == "FFFFEB9C"
    assert sheet["F3"].fill.fgColor.rgb == "FFC6EFCE"
    assert sheet["G3"].fill.fgColor.rgb == "FFFFC7CE"
    assert sheet["H3"].fill.fgColor.rgb == "FFFFC7CE"
    assert sheet["I3"].fill.fgColor.rgb == "FFFFC7CE"
    assert sheet["A2"].alignment.wrap_text is True
    assert sheet["E3"].alignment.wrap_text is not True
    assert sheet["G3"].alignment.wrap_text is not True


def test_export_writes_one_safely_named_workbook_per_selected_product_line(tmp_path):
    lines = (
        ProductLine("DOPL", "https://c/dopl", "Digital"),
        ProductLine("SDPL", "https://c/sdpl", "Set Top Box"),
        ProductLine("TV", "https://c/tv", "TV / Video"),
        ProductLine("OOPL", "https://c/oopl", "Operator"),
    )
    batch = AuditBatch(
        "batch1", AuditPeriod(
            datetime(2026, 8, 3, tzinfo=TZ),
            datetime(2026, 8, 6, tzinfo=TZ),
        ),
        datetime(2026, 8, 6, tzinfo=TZ),
        product_lines=lines,
    )

    paths = export_project_audit_xlsx_by_product_line(batch, tmp_path)

    assert [path.name for path in paths] == [
        "project_weekly_audit_Digital_batch1.xlsx",
        "project_weekly_audit_Set_Top_Box_batch1.xlsx",
        "project_weekly_audit_TV_Video_batch1.xlsx",
        "project_weekly_audit_Operator_batch1.xlsx",
    ]
    assert all(path.is_file() for path in paths)
    for path in paths:
        sheet = load_workbook(path).active
        assert [cell.value for cell in sheet[2]] == [
            "Owner", "年份", "项目名", "项目链接",
            *(point.label for point in UPDATE_MATRIX_POINTS), "PS",
        ]
        assert sheet.max_row == 2


def test_product_line_export_with_zero_selected_lines_writes_nothing(tmp_path):
    batch = AuditBatch(
        "batch0", AuditPeriod(
            datetime(2026, 8, 3, tzinfo=TZ),
            datetime(2026, 8, 6, tzinfo=TZ),
        ),
        datetime(2026, 8, 6, tzinfo=TZ),
    )

    assert export_project_audit_xlsx_by_product_line(batch, tmp_path) == []
    assert list(tmp_path.iterdir()) == []
