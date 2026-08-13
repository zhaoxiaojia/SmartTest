from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from tool.common.jira_format_audit import (
    AuditReport,
    AuditViolation,
    IssueAuditResult,
    JiraAuditService,
    ResolvedAuditInput,
    active_rules,
    audit_issue,
    export_audit_xlsx,
    resolve_audit_input,
)
from tool.common.jira_format_audit.rules import (
    ai_reviewable_violations,
    is_audit_eligible,
)
from support.jira_integration.core.models import SearchPage


RULE_IDS = tuple(
    """
    SUMMARY.FORMAT COMPONENT.REQUIRED
    DESCRIPTION.STEPS_TO_REPRODUCE DESCRIPTION.ACTUAL_RESULTS
    DESCRIPTION.EXPECTED_RESULTS DESCRIPTION.COMPARISON DESCRIPTION.NOTES
    DESCRIPTION.RATE_FORMAT DESCRIPTION.NOTES_HW DESCRIPTION.NOTES_SW
    DESCRIPTION.TABLE_REQUIRED_VALUE
    """.split()
)
_DEFAULT_CREATOR = object()


def _description(
    *,
    steps="1. Start playback.\n2. Seek to 00:30.",
    actual="Video freezes.",
    expected="Playback continues.",
    rate="2/2",
    comparison="Previous version V1.0 is normal; current version V1.1 is broken.",
    notes="HW info: T7 reference board\nSW info: V1.1",
):
    values = (
        ("Steps to reproduce", steps),
        ("Actual results", actual),
        ("Expected results", expected),
        ("Reproducibility rate", rate),
        ("Comparison", comparison),
        ("Notes", notes),
    )
    return "\n".join(f"[{heading}]:\n{value}" for heading, value in values)


def _table_description(*rows):
    values = (
        ("平台信息", "客户/项目代号", "T6X高刷"),
        ("", "主芯片", "T966D5"),
        ("测试环境", "测试仪器", "-"),
        ("测试设置", "HDMI输出", "无"),
        *rows,
    )
    return "\n".join(
        (
            "||模块||需要填写信息||测试信息||",
            *("|" + "|".join(row) + "|" for row in values),
        )
    )


def _issue(
    key="SH-123",
    *,
    summary="[ACME][T7][V1.1][Video]: Video freezes after seeking,2/2",
    description=None,
    components=("Video",),
    labels=("regression",),
    creator=_DEFAULT_CREATOR,
):
    return {
        "key": key,
        "fields": {
            "summary": summary,
            "description": _description() if description is None else description,
            "reporter": {"displayName": "Coco"},
            "creator": (
                {"displayName": "Chao Li"}
                if creator is _DEFAULT_CREATOR
                else creator
            ),
            "components": [{"name": value} for value in components],
            "labels": list(labels),
        },
    }


def _violations(**changes):
    return {
        item.rule_id
        for item in audit_issue(_issue(**changes), base_url="https://jira.example.com").violations
    }


def test_complete_issue_accepts_four_five_and_six_summary_groups():
    summaries = (
        "[ACME][T7][V1.1][Video]: Video freezes,50%",
        "[SH-1][ACME][T7][V1.1][Video]: Video freezes,1/2",
        "[SH-1][BUG-2][ACME][T7][V1.1][Video]: Video freezes,100%.",
        "【ACME】 [T7] 【V1.1】 [Video]： Video freezes,50%",
    )

    assert tuple(rule.rule_id for rule in active_rules()) == RULE_IDS
    assert all(audit_issue(_issue(summary=value), base_url="https://jira.example.com").passed
               for value in summaries)


@pytest.mark.parametrize(
    "suffix",
    (
        "",
        " (10%)",
        "（10%）",
        "；1/6台",
        "，偶发且次数暂未统计",
    ),
)
def test_summary_probability_text_is_not_audited(suffix):
    summary = f"[ACME][T7][V1.1][Video]: Video freezes{suffix}"

    assert not _violations(summary=summary)


@pytest.mark.parametrize("rate", ("10%", "1/2", "出现一次"))
def test_summary_with_only_a_standard_probability_has_no_problem_description(rate):
    summary = f"[ACME][T7][V1.1][Video]: {rate}"

    assert _violations(summary=summary) == {"SUMMARY.FORMAT"}


def test_nonempty_ambiguous_summary_format_is_ai_reviewable():
    summary = "[ACME][T7][V1.1][Video] Video freezes on 1/6台"
    result = audit_issue(
        _issue(summary=summary),
        base_url="https://jira.example.com",
    )

    assert [item.rule_id for item in result.violations] == ["SUMMARY.FORMAT"]
    assert [item.rule_id for item in ai_reviewable_violations(result)] == [
        "SUMMARY.FORMAT"
    ]


@pytest.mark.parametrize("summary", ("", "invalid"))
def test_empty_or_uninformative_summary_format_is_not_ai_reviewable(summary):
    result = audit_issue(
        _issue(summary=summary),
        base_url="https://jira.example.com",
    )

    assert "SUMMARY.FORMAT" in {
        item.rule_id for item in result.violations
    }
    assert not ai_reviewable_violations(result)


@pytest.mark.parametrize(
    "rate",
    (
        "50%",
        "1/2",
        "出现一次",
        "复现2次",
        "出现三次",
        "出现一次（问题出现后一直不能恢复）",
        "复现2次(problem remains)",
    ),
)
def test_description_accepts_supported_occurrence_rates(rate):
    assert "DESCRIPTION.RATE_FORMAT" not in _violations(
        description=_description(rate=rate)
    )


@pytest.mark.parametrize(
    "rate",
    ("偶现", "出现2", "出现一次 problem remains"),
)
def test_description_rejects_ambiguous_or_unbounded_rate_text(rate):
    assert "DESCRIPTION.RATE_FORMAT" in _violations(
        description=_description(rate=rate)
    )


def test_description_keeps_hw_and_sw_after_nested_expected_results_in_notes():
    description = _description(
        notes=(
            "Test Case:\n1. Play a video.\nExpected results:\nPlayback continues.\n"
            "HW info: DVT2\nSW info: V460"
        )
    )

    violations = _violations(description=description)

    assert "DESCRIPTION.NOTES_HW" not in violations
    assert "DESCRIPTION.NOTES_SW" not in violations


def test_description_accepts_unbracketed_chinese_semantic_headings():
    description = """操作步骤：
1. 打开电视并播放视频。
实际结果：播放后画面冻结。
期望结果：视频应持续播放。
概率：2/2
对比信息：上一版本正常，当前版本异常。
备注：现场复现信息如下。
软件版本：V460
硬件信息：DVT2
"""

    assert not _violations(description=description)


def test_tv_style_wrapped_chinese_headings_and_supplemental_info_are_parsed():
    description = """【操作步骤】：
1. 打开电视并播放视频。
【实际结果】；
播放后画面冻结。
【期望结果】：
视频应持续播放。
【概率】：100%
【软件版本】：V460
【硬件信息】：65寸
"""

    assert _violations(description=description) == {
        "DESCRIPTION.COMPARISON"
    }


def test_all_supported_english_headings_work_without_brackets():
    description = _description().replace("[", "").replace("]", "")

    assert not _violations(description=description)


def test_second_description_template_accepts_nonempty_third_column_values():
    result = audit_issue(
        _issue(description=_table_description()),
        base_url="https://jira.example.com",
    )

    assert result.passed


def test_second_description_template_reports_every_empty_third_column_row():
    result = audit_issue(
        _issue(
            description=_table_description(
                ("测试环境", "测试地点", ""),
                ("", "测试人员/日期", "   "),
            )
        ),
        base_url="https://jira.example.com",
    )

    violations = [
        item
        for item in result.violations
        if item.rule_id == "DESCRIPTION.TABLE_REQUIRED_VALUE"
    ]
    assert [item.field for item in violations] == [
        "Description.测试信息",
        "Description.测试信息",
    ]
    assert [item.reason for item in violations] == [
        "Description 表格中“测试地点”行的测试信息为空。",
        "Description 表格中“测试人员/日期”行的测试信息为空。",
    ]


def test_non_table_description_still_uses_first_description_template_rules():
    assert _violations(description=_description(actual="")) == {
        "DESCRIPTION.ACTUAL_RESULTS"
    }


EMPTY_DESCRIPTION = _description(
    steps="",
    actual="",
    expected="",
    rate="",
    comparison="",
    notes="",
)


@pytest.mark.parametrize(
    ("changes", "expected"),
    (
        ({"summary": "invalid"}, {"SUMMARY.FORMAT"}),
        (
            {"summary": "[ ][ ][ ][Video]: Video freezes,50%"},
            {"SUMMARY.FORMAT"},
        ),
        ({"summary": "[客户][T7][V1][Video]: 播放冻结,50%"}, set()),
        ({"summary": "[ACME][T7][V1][Video]: Video freezes,3/2"}, set()),
        ({"components": ()}, {"COMPONENT.REQUIRED"}),
        (
            {"description": EMPTY_DESCRIPTION, "labels": ()},
            {
                "DESCRIPTION.STEPS_TO_REPRODUCE",
                "DESCRIPTION.ACTUAL_RESULTS",
                "DESCRIPTION.EXPECTED_RESULTS",
                "DESCRIPTION.COMPARISON",
                "DESCRIPTION.NOTES",
                "DESCRIPTION.NOTES_HW",
                "DESCRIPTION.NOTES_SW",
            },
        ),
        (
            {
                "description": _description(
                    rate="3/2",
                    notes="HW info:\nSW info:",
                ),
                "labels": (),
            },
            {
                "DESCRIPTION.RATE_FORMAT",
                "DESCRIPTION.NOTES_HW",
                "DESCRIPTION.NOTES_SW",
            },
        ),
        ({"description": _description(comparison="V1 and V2 compared.")}, set()),
    ),
)
def test_rule_failure_matrix(changes, expected):
    assert expected == _violations(**changes)


@pytest.mark.parametrize(
    ("rule_id", "changes", "source_field"),
    (
        ("SUMMARY.FORMAT", {"summary": "invalid summary"}, "summary"),
        ("COMPONENT.REQUIRED", {"components": ()}, None),
        (
            "DESCRIPTION.STEPS_TO_REPRODUCE",
            {"description": _description(steps="")},
            "description",
        ),
        (
            "DESCRIPTION.ACTUAL_RESULTS",
            {"description": _description(actual="")},
            "description",
        ),
        (
            "DESCRIPTION.EXPECTED_RESULTS",
            {"description": _description(expected="")},
            "description",
        ),
        (
            "DESCRIPTION.COMPARISON",
            {"description": _description(comparison="")},
            "description",
        ),
        (
            "DESCRIPTION.NOTES",
            {"description": _description(notes="")},
            "description",
        ),
        (
            "DESCRIPTION.RATE_FORMAT",
            {"description": _description(rate="often")},
            "description",
        ),
        (
            "DESCRIPTION.NOTES_HW",
            {"description": _description(notes="SW info: V1.1")},
            "description",
        ),
        (
            "DESCRIPTION.NOTES_SW",
            {"description": _description(notes="HW info: T7")},
            "description",
        ),
    ),
)
def test_each_violation_keeps_the_full_original_jira_field(
    rule_id, changes, source_field
):
    issue = _issue(**changes)
    result = audit_issue(issue, base_url="https://jira.example.com")
    violation = next(
        item for item in result.violations if item.rule_id == rule_id
    )
    expected = "[]" if source_field is None else issue["fields"][source_field]

    assert violation.observed == expected
    assert violation.observed


def test_legacy_comparision_heading_remains_violation_but_uses_semantic_review():
    description = _description().replace("[Comparison]", "[Comparision]")

    result = audit_issue(
        _issue(description=description, labels=()),
        base_url="https://jira.example.com",
    )

    assert {item.rule_id for item in result.violations} == {
        "DESCRIPTION.COMPARISON"
    }
    assert [item.rule_id for item in ai_reviewable_violations(result)] == [
        "DESCRIPTION.COMPARISON"
    ]


def test_only_declared_ambiguous_violations_become_ai_candidates():
    hard_failure = audit_issue(
        _issue(description=_description(actual="")),
        base_url="https://jira.example.com",
    )
    fuzzy_failures = (
        audit_issue(
            _issue(
                summary="[ACME][T7][V1.1][Video] Video freezes",
                labels=(),
            ),
            base_url="https://jira.example.com",
        ),
        audit_issue(
            _issue(description=_description(rate="intermittent"), labels=()),
            base_url="https://jira.example.com",
        ),
        audit_issue(
            _issue(description=_description(comparison=""), labels=()),
            base_url="https://jira.example.com",
        ),
        audit_issue(
            _issue(
                description=_description(
                    notes="Hardware information:\nSoftware information: V1.1"
                ),
                labels=(),
            ),
            base_url="https://jira.example.com",
        ),
        audit_issue(
            _issue(
                description=_description(notes="HW info: T7\nSW info:"),
                labels=(),
            ),
            base_url="https://jira.example.com",
        ),
    )

    assert not ai_reviewable_violations(hard_failure)
    assert [bool(ai_reviewable_violations(result)) for result in fuzzy_failures] == [
        True,
        True,
        True,
        True,
        True,
    ]
    full_description = _description(
        actual="HW info: T7 reference board\nVideo freezes.",
        notes="SW info: V1.1",
    )
    full_context_result = audit_issue(
        _issue(description=full_description, labels=()),
        base_url="https://jira.example.com",
    )
    assert not ai_reviewable_violations(
        full_context_result,
        description=full_description,
    )
    assert "DESCRIPTION.NOTES_HW" not in {
        item.rule_id for item in full_context_result.violations
    }


def test_unlabeled_natural_language_fields_become_ai_candidates():
    description = """打开电视后进入播放器并播放视频。
播放过程中画面冻结。
视频应该持续播放而不冻结。
该问题偶尔出现。
上一版本播放正常，当前版本异常。
补充说明：使用 DVT2 机型，测试版本为 V460。
"""
    result = audit_issue(
        _issue(description=description, labels=()),
        base_url="https://jira.example.com",
    )

    assert {item.rule_id for item in ai_reviewable_violations(
        result,
        description=description,
    )} == {
        "DESCRIPTION.STEPS_TO_REPRODUCE",
        "DESCRIPTION.ACTUAL_RESULTS",
        "DESCRIPTION.EXPECTED_RESULTS",
        "DESCRIPTION.COMPARISON",
        "DESCRIPTION.NOTES",
        "DESCRIPTION.NOTES_HW",
        "DESCRIPTION.NOTES_SW",
    }


def test_unrelated_text_does_not_make_missing_sections_ai_candidates():
    description = "Video playback issue."
    result = audit_issue(
        _issue(description=description, labels=()),
        base_url="https://jira.example.com",
    )

    assert not ai_reviewable_violations(result, description=description)


def test_active_rule_text_does_not_reintroduce_english_only_semantics():
    rules = {rule.rule_id: rule for rule in active_rules()}
    summary_format = rules["SUMMARY.FORMAT"]
    description_rate = rules["DESCRIPTION.RATE_FORMAT"]

    assert summary_format.requirement == (
        "Summary 必须包含 4–6 个方括号分组，最后四组依次为客户、CHIP、"
        "系统版本和模块，冒号后填写非空问题描述。"
    )
    assert summary_format.guidance == (
        "使用“[客户][CHIP][版本][模块]: 问题描述”；"
        "前面可选增加公共 Jira ID 和客户 Bug ID。"
    )
    assert all(
        "英文" not in rule.requirement + rule.guidance
        and "English" not in rule.requirement + rule.guidance
        for rule in rules.values()
    )
    assert "SUMMARY.PROBABILITY" not in rules
    assert "百分比、分数或明确的文字次数" in description_rate.requirement
    assert "出现一次" in (
        description_rate.requirement + description_rate.guidance
    )
    assert "括号" in description_rate.guidance


@pytest.mark.parametrize(
    ("creator", "expected"),
    (
        ({"displayName": "  CHAO   LI  "}, True),
        ({"name": r"AMLOGIC\chao.li"}, True),
        ({"key": "chao.li@example.com"}, True),
        ({"accountId": "chao.li"}, True),
        ("Chao Li", True),
        ({"displayName": "Someone Else", "name": "chao.li"}, False),
        ({"displayName": "Someone Else"}, False),
        (None, False),
    ),
)
def test_creator_eligibility_uses_the_smarttest_allowlist(creator, expected):
    assert is_audit_eligible(_issue(creator=creator)) is expected


@dataclass
class FakeClient:
    pages: list[SearchPage] | None = None
    filter_payload: dict | None = None
    error: Exception | None = None

    def __post_init__(self):
        self.search_calls = []
        self.filter_calls = []

    def search_page(self, jql, **kwargs):
        self.search_calls.append((jql, kwargs))
        if self.error:
            raise self.error
        if self.pages:
            return self.pages.pop(0)
        return SearchPage([], 0, kwargs.get("max_results", 1), 0, True)

    def fetch_filter(self, filter_id):
        self.filter_calls.append(filter_id)
        return dict(self.filter_payload or {})


@pytest.mark.parametrize(
    ("text", "kind", "jql", "filter_payload"),
    (
        ("project = SH", "jql", "project = SH", None),
        ("https://jira.example.com/browse/SH-123", "issue_url", 'key = "SH-123"', None),
        ("https://jira.example.com/issues/?jql=project%20%3D%20SH",
         "jql_url", "project = SH", None),
        ("https://jira.example.com/filter/42", "filter_url", "labels = regression",
         {"jql": "labels = regression"}),
    ),
)
def test_resolves_supported_jql_and_jira_urls(text, kind, jql, filter_payload):
    client = FakeClient(filter_payload=filter_payload)

    result = resolve_audit_input(text, base_url="https://jira.example.com", client=client)

    assert (result.source_kind, result.jql) == (kind, jql)
    assert client.search_calls[0][0] == jql
    assert client.search_calls[0][1]["fields"] == (
        "creator",
        "summary",
        "description",
        "reporter",
        "components",
    )


def test_rejects_empty_external_or_invalid_jira_input():
    invalid = (
        ("", "JQL"),
        ("https://other.example.com/browse/SH-1", "host"),
        ("ftp://jira.example.com/browse/SH-1", "HTTP"),
        ("https://jira.example.com/browse/not-a-key", "URL"),
    )
    for text, message in invalid:
        with pytest.raises(ValueError, match=message):
            resolve_audit_input(text, base_url="https://jira.example.com", client=FakeClient())
    with pytest.raises(ValueError, match="JQL"):
        resolve_audit_input("project =", base_url="https://jira.example.com",
                            client=FakeClient(error=RuntimeError("invalid")))


def test_service_paginates_and_reports_stable_progress_stages():
    client = FakeClient(
        pages=[
            SearchPage([_issue("SH-1"), _issue("SH-2")], 0, 2, 5, False),
            SearchPage([_issue("SH-2"), _issue("SH-3")], 0, 2, 5, False),
            SearchPage([_issue("SH-4")], 4, 2, 5, True),
        ]
    )
    progress = []

    report = JiraAuditService(client, base_url="https://jira.example.com").run(
        ResolvedAuditInput("jql", "project = SH", "project = SH"),
        lambda *value: progress.append(value),
    )

    assert [call[1]["start_at"] for call in client.search_calls] == [0, 2, 4]
    assert [item.key for item in report.issues] == ["SH-1", "SH-2", "SH-3", "SH-4"]
    assert report.passed_count == 4
    assert progress == [
        ("fetching", 2, 5),
        ("fetching", 4, 5),
        ("fetching", 5, 5),
        ("rule_auditing", 4, 4),
        ("ai_reviewing", 0, 0),
        ("finalizing", 4, 4),
    ]


def test_empty_service_run_reports_every_stable_progress_stage():
    progress = []

    JiraAuditService(FakeClient(), base_url="https://jira.example.com").run(
        ResolvedAuditInput("jql", "project = NONE", "project = NONE"),
        lambda *value: progress.append(value),
    )

    assert progress == [
        ("fetching", 0, 0),
        ("rule_auditing", 0, 0),
        ("ai_reviewing", 0, 0),
        ("finalizing", 0, 0),
    ]


def test_openpyxl_export_preserves_reader_visible_report_contract(tmp_path):
    summary = "[ACME][T7][V1.1][Video]: Video freezes,often"
    dynamic_reason = "Summary probability is invalid."
    dynamic_guidance = "Use a percentage or fraction."
    rules_by_id = {rule.rule_id: rule for rule in active_rules()}
    format_rule = rules_by_id["SUMMARY.FORMAT"]
    format_violation = AuditViolation(
        format_rule.rule_id,
        format_rule.section,
        format_rule.field,
        summary,
        "Summary grouping is invalid.",
        format_rule.guidance,
    )
    component_violation = AuditViolation(
        "COMPONENT.REQUIRED",
        "Component",
        "Component",
        "[]",
        "This supplied reason must be replaced by report text.",
        "This supplied guidance must be replaced by report text.",
    )
    rate_violation = AuditViolation(
        "DESCRIPTION.RATE_FORMAT",
        "Description",
        "Description.Reproducibility rate",
        summary,
        dynamic_reason,
        dynamic_guidance,
    )
    issues = (
        IssueAuditResult(
            "SH-2",
            "https://jira.example.com/browse/SH-2",
            summary,
            "Bob",
            False,
            (component_violation, rate_violation, format_violation),
        ),
        IssueAuditResult(
            "SH-1",
            "https://jira.example.com/browse/SH-1",
            "Valid summary",
            "Bob",
            False,
            (component_violation,),
        ),
        IssueAuditResult(
            "SH-3",
            "https://jira.example.com/browse/SH-3",
            "Valid summary",
            "Alice",
            True,
            (),
        ),
    )
    generated_at = datetime(2026, 7, 24, 10, 11, 12, tzinfo=timezone.utc)
    jql = "=project = SH"
    report = AuditReport(
        resolved=ResolvedAuditInput("jql", jql, jql),
        generated_at=generated_at,
        rules=active_rules(),
        issues=issues,
    )
    smart_dir = tmp_path / "smart"
    first = export_audit_xlsx(report, downloads_dir=smart_dir, now=generated_at)
    second = export_audit_xlsx(report, downloads_dir=smart_dir, now=generated_at)
    smart_workbook = load_workbook(first)
    assert smart_workbook.sheetnames == ["汇总", "违规明细"]
    summary_sheet = smart_workbook["汇总"]
    detail_sheet = smart_workbook["违规明细"]
    assert tuple(summary_sheet.values)[:10] == (
        ("指标", "值", None),
        ("生成时间", generated_at.isoformat(), None),
        ("JQL 查询条件", jql, None),
        ("问题总数", 3, None),
        ("通过 Jira 数", 1, None),
        ("不通过 Jira 数", 2, None),
        ("通过率", "33.33%", None),
        (None, None, None),
        ("报告人", "违规 Jira 数量", "违规 Jira 号"),
        ("Bob", 2, "SH-1、SH-2"),
    )
    assert detail_sheet.freeze_panes == summary_sheet.freeze_panes == "A2"
    assert detail_sheet.auto_filter.ref == "A1:J5"
    assert {
        str(value) for value in detail_sheet.merged_cells.ranges
    } == {"A2:A4", "B2:B4", "C2:C4"}
    assert detail_sheet["A1"].font.bold
    assert not detail_sheet["A2"].font.bold
    assert first.name == "jira_format_audit_20260724_101112.xlsx"
    assert second.name == "jira_format_audit_20260724_101112_2.xlsx"
    assert not list(smart_dir.glob("*.tmp"))
    assert smart_workbook["汇总"]["B3"].value == jql
    assert smart_workbook["汇总"]["B3"].data_type == "s"
    detail_text = " ".join(
        str(cell.value or "")
        for row in smart_workbook["违规明细"].iter_rows()
        for cell in row
    )
    assert "REGRESSION.EVIDENCE" not in detail_text
    assert "英文问题描述" not in detail_text
