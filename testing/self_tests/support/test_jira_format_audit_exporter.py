from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]


def _report():
    from support.jira_integration.audit import (
        AuditReport,
        AuditViolation,
        IssueAuditResult,
        ResolvedAuditInput,
        active_rules,
    )

    violation = AuditViolation(
        rule_id="SUMMARY-001",
        section="Summary",
        field="summary",
        observed="bad summary",
        reason="Summary 结构不符合规范。",
        guidance="按规则修改。",
    )
    issue = IssueAuditResult(
        key="SH-1",
        url="https://jira.example.com/browse/SH-1",
        summary="bad summary",
        reporter="Coco",
        passed=False,
        violations=(violation,),
    )
    return AuditReport(
        resolved=ResolvedAuditInput("jql", "project = SH", "project = SH"),
        generated_at=datetime(2026, 7, 24, 10, 11, 12),
        rules=active_rules(),
        issues=(issue,),
    )


def test_export_creates_unique_openpyxl_workbook_with_four_sheets_and_hyperlinks(
    tmp_path,
):
    from support.jira_integration.audit.exporter import export_audit_xlsx

    now = datetime(2026, 7, 24, 10, 11, 12)
    first = export_audit_xlsx(_report(), downloads_dir=tmp_path, now=now)
    second = export_audit_xlsx(_report(), downloads_dir=tmp_path, now=now)

    assert first.name == "jira_format_audit_20260724_101112.xlsx"
    assert second.name == "jira_format_audit_20260724_101112_2.xlsx"
    assert first.exists() and second.exists()
    assert not list(tmp_path.glob("*.tmp"))

    workbook = load_workbook(first, data_only=False)
    assert workbook.sheetnames == ["Summary", "Rules", "Issues", "Violations"]
    assert workbook["Summary"]["B5"].value == "project = SH"
    assert workbook["Rules"]["A2"].value == "SUMMARY.FORMAT"
    assert workbook["Issues"]["A2"].value == "SH-1"
    assert workbook["Issues"]["B2"].value == "https://jira.example.com/browse/SH-1"
    assert workbook["Issues"]["B2"].hyperlink.target == "https://jira.example.com/browse/SH-1"
    assert workbook["Violations"]["C2"].value == "SUMMARY-001"
    assert workbook["Violations"]["H2"].value == "Summary 结构不符合规范。"
    assert workbook["Violations"]["B2"].hyperlink.target == (
        "https://jira.example.com/browse/SH-1"
    )
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert all(cell.font.bold for cell in sheet[1])


def test_exporter_reuses_declared_openpyxl_without_manual_ooxml():
    exporter = (ROOT / "support/jira_integration/audit/exporter.py").read_text(
        encoding="utf-8"
    )
    dependencies = (ROOT / "support/scripts/script-init-venv.py").read_text(
        encoding="utf-8"
    )
    pyinstaller = (ROOT / "support/packaging/pyinstaller/main.spec").read_text(
        encoding="utf-8"
    )

    assert "from openpyxl import Workbook" in exporter
    assert "openpyxl==3.1.5" in dependencies
    for obsolete in (
        "ZipFile",
        "zipfile",
        "_sheet_xml",
        "_content_types",
        "_package_relationships",
        "_workbook_xml",
        "_workbook_relationships",
        "_styles_xml",
    ):
        assert obsolete not in exporter
    assert "openpyxl" not in pyinstaller
