from __future__ import annotations

import importlib
import json
import re
import weakref
from dataclasses import asdict
from threading import Barrier, Lock

import pytest
from openpyxl import load_workbook

from support.ai import (
    AIChatResponse,
    AIConfigurationError,
    AIResponseError,
    AITransportError,
)
from support.jira_integration.audit import (
    AIReviewStatus,
    JiraAuditService,
    ResolvedAuditInput,
    export_audit_xlsx,
)
from support.jira_integration.core.models import SearchPage


service_module = importlib.import_module("support.jira_integration.audit.service")
_DEFAULT_CREATOR = object()


def _description(
    *,
    actual="Video freezes.",
    rate="2/2",
    comparison="Previous version V1.0 is normal; current version V1.1 is broken.",
    notes="HW info: T7 reference board\nSW info: V1.1",
):
    values = (
        ("Steps to reproduce", "1. Start playback.\n2. Seek to 00:30."),
        ("Actual results", actual),
        ("Expected results", "Playback continues."),
        ("Reproducibility rate", rate),
        ("Comparison", comparison),
        ("Notes", notes),
    )
    return "\n".join(f"[{heading}]:\n{value}" for heading, value in values)


def _issue(
    key="SH-123",
    *,
    summary="[ACME][T7][V1.1][Video]: Video freezes after seeking,2/2",
    description=None,
    creator=_DEFAULT_CREATOR,
):
    return {
        "key": key,
        "fields": {
            "summary": summary,
            "description": _description() if description is None else description,
            "reporter": {"displayName": "Coco"},
            "creator": (
                {"displayName": "Chao Li"}
                if creator is _DEFAULT_CREATOR
                else creator
            ),
            "components": [{"name": "Video"}],
            "labels": [],
            "attachment": [],
        },
    }


class JiraClient:
    def __init__(self, issues):
        self._issues = list(issues)

    def search_page(self, _jql, **kwargs):
        return SearchPage(
            self._issues,
            kwargs.get("start_at", 0),
            len(self._issues),
            len(self._issues),
            True,
        )


class AIClient:
    def __init__(self, responses):
        self.responses = dict(responses)
        self.requests = []

    def chat_completion(self, messages, **kwargs):
        self.requests.append((messages, kwargs))
        content = messages[-1].content
        issue_key = re.search(r'"issue_key":\s*"([^"]+)"', content).group(1)
        response = self.responses[issue_key]
        if isinstance(response, BaseException):
            raise response
        return AIChatResponse(content=response, model="test-model")


class ConcurrentAIClient(AIClient):
    def __init__(self, responses):
        super().__init__(responses)
        self._barrier = Barrier(6)
        self._lock = Lock()
        self._started = 0
        self._active = 0
        self.max_active = 0

    def chat_completion(self, messages, **kwargs):
        with self._lock:
            self._started += 1
            self._active += 1
            self.max_active = max(self.max_active, self._active)
            wait_for_peers = self._started <= 6
        try:
            if wait_for_peers:
                self._barrier.wait(timeout=1)
            return super().chat_completion(messages, **kwargs)
        finally:
            with self._lock:
                self._active -= 1


def _run(monkeypatch, issues, factory, progress=None):
    monkeypatch.setattr(service_module, "create_chat_client", factory)
    return JiraAuditService(JiraClient(issues), base_url="https://jira.example.com").run(
        ResolvedAuditInput("jql", "project = SH", "project = SH"),
        progress or (lambda *_value: None),
    )


def _review(issue_key, *decisions):
    return json.dumps(
        {
            "issue_key": issue_key,
            "decisions": [
                {
                    "rule_id": rule_id,
                    "result": result,
                    "reason": reason,
                    "guidance": guidance,
                }
                for rule_id, result, reason, guidance in decisions
            ],
        }
    )


def test_one_jira_with_multiple_candidates_uses_one_ai_request_and_merges_results(
    monkeypatch,
):
    client = AIClient(
        {
            "SH-1": _review(
                "SH-1",
                ("SUMMARY.PROBABILITY", "PASS", "", ""),
                (
                    "DESCRIPTION.RATE_FORMAT",
                    "FAIL",
                    "原文没有明确量化复现概率。",
                    "补充百分比或分数。",
                ),
            )
        }
    )
    issue = _issue(
        "SH-1",
        summary="[ACME][T7][V1.1][Video]: Video freezes,often",
        description=_description(rate="intermittent"),
    )

    result = _run(monkeypatch, [issue], lambda: client).issues[0]

    assert len(client.requests) == 1
    assert result.ai_review_status is AIReviewStatus.COMPLETED
    assert [item.rule_id for item in result.violations] == [
        "DESCRIPTION.RATE_FORMAT"
    ]
    assert result.violations[0].reason == "原文没有明确量化复现概率。"
    assert result.violations[0].guidance == "补充百分比或分数。"


def test_ai_receives_full_jira_context(monkeypatch):
    description_marker = "private-description-context"
    description = _description(
        actual=f"Video freezes; {description_marker}; observed 2/2.",
        rate="intermittent",
    )
    summary = "[ACME][T7][V1.1][Video]: Video freezes,intermittent"
    client = AIClient(
        {
            "SH-1": json.dumps(
                {
                    "issue_key": "SH-1",
                    "decisions": [
                        {
                            "rule_id": "SUMMARY.PROBABILITY",
                            "result": "PASS",
                            "reason": "",
                            "guidance": "",
                        },
                        {
                            "rule_id": "DESCRIPTION.RATE_FORMAT",
                            "result": "PASS",
                            "reason": "",
                            "guidance": "",
                        },
                    ],
                }
            )
        }
    )

    report = _run(
        monkeypatch,
        [_issue("SH-1", summary=summary, description=description)],
        lambda: client,
    )
    request_payload = json.loads(
        client.requests[0][0][-1].content.rsplit("\n", 1)[-1]
    )
    assert request_payload["jira_fields"] == {
        "Summary": summary,
        "Description": description,
    }
    assert "description" not in asdict(report.issues[0])


def test_report_and_workbook_retain_description_only_as_violation_observed(
    monkeypatch, tmp_path
):
    description_marker = "private-description-context"
    description = _description(actual="") + f"\n{description_marker}"
    report = _run(
        monkeypatch,
        [_issue("SH-1", description=description)],
        lambda: pytest.fail("hard violations must not create an AI client"),
    )

    assert "description" not in asdict(report.issues[0])
    assert report.issues[0].violations[0].observed == description

    exported = export_audit_xlsx(report, downloads_dir=tmp_path)
    workbook = load_workbook(exported)
    workbook_text = "\n".join(
        str(cell.value or "")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    assert description_marker in workbook_text


def test_hard_violation_does_not_create_ai_client(monkeypatch):
    created = []
    issue = _issue("SH-1", description=_description(actual=""))

    report = _run(monkeypatch, [issue], lambda: created.append(True))

    assert not created
    assert report.issues[0].ai_review_status is AIReviewStatus.NOT_REQUIRED
    assert {item.rule_id for item in report.issues[0].violations} == {
        "DESCRIPTION.ACTUAL_RESULTS"
    }


def test_unconfigured_ai_retains_initial_violation_with_sanitized_status(monkeypatch):
    issue = _issue(
        "SH-1",
        summary="[ACME][T7][V1.1][Video]: Video freezes,often",
    )

    result = _run(
        monkeypatch,
        [issue],
        lambda: (_ for _ in ()).throw(
            AIConfigurationError("secret key must not be reported")
        ),
    ).issues[0]

    assert result.ai_review_status is AIReviewStatus.UNCONFIGURED
    assert [item.rule_id for item in result.violations] == [
        "SUMMARY.PROBABILITY"
    ]
    assert "secret" not in repr(result)


@pytest.mark.parametrize(
    "failure",
    (
        TimeoutError("private timeout detail"),
        AITransportError(
            "private transport detail", category="connectionreseterror"
        ),
        AIResponseError("private response detail"),
    ),
)
def test_ai_failure_retains_initial_violation_and_sanitized_status(
    monkeypatch, failure
):
    issue = _issue(
        "SH-1",
        summary="[ACME][T7][V1.1][Video]: Video freezes,often",
    )
    client = AIClient({"SH-1": failure})

    result = _run(monkeypatch, [issue], lambda: client).issues[0]

    assert result.ai_review_status is AIReviewStatus.FAILED
    assert [item.rule_id for item in result.violations] == [
        "SUMMARY.PROBABILITY"
    ]
    assert "private" not in repr(result)


@pytest.mark.parametrize(
    "response",
    (
        "",
        "{}",
        _review(
            "OTHER-1",
            ("SUMMARY.PROBABILITY", "PASS", "", ""),
        ),
        _review("SH-1"),
        _review(
            "SH-1",
            ("SUMMARY.PROBABILITY", "PASS", "", ""),
            ("SUMMARY.PROBABILITY", "PASS", "", ""),
        ),
        _review(
            "SH-1",
            ("SUMMARY.PROBABILITY", "MAYBE", "", ""),
        ),
        _review(
            "SH-1",
            ("SUMMARY.PROBABILITY", "FAIL", "", "补充概率。"),
        ),
    ),
)
def test_invalid_ai_decision_payload_degrades_without_changing_violations(
    monkeypatch, response
):
    issue = _issue(
        "SH-1",
        summary="[ACME][T7][V1.1][Video]: Video freezes,often",
    )

    result = _run(
        monkeypatch, [issue], lambda: AIClient({"SH-1": response})
    ).issues[0]

    assert result.ai_review_status is AIReviewStatus.FAILED
    assert [item.rule_id for item in result.violations] == [
        "SUMMARY.PROBABILITY"
    ]


def test_ai_reviews_use_six_workers_preserve_order_and_isolate_failures(monkeypatch):
    keys = [f"SH-{index}" for index in range(1, 8)]
    responses = {
        key: _review(key, ("SUMMARY.PROBABILITY", "PASS", "", ""))
        for key in keys
    }
    responses["SH-1"] = "{}"
    client = ConcurrentAIClient(responses)
    issues = [
        _issue(
            key,
            summary=f"[ACME][T7][V1.1][Video]: Video freezes {key},often",
        )
        for key in keys
    ]
    progress = []
    report = _run(
        monkeypatch,
        issues,
        lambda: client,
        lambda *value: progress.append(value),
    )

    assert client.max_active == 6
    assert [item.key for item in report.issues] == keys
    assert [item.ai_review_status for item in report.issues] == [
        AIReviewStatus.FAILED,
        *([AIReviewStatus.COMPLETED] * 6),
    ]
    assert [item.passed for item in report.issues] == [False, *([True] * 6)]
    assert [
        value[1:] for value in progress if value[0] == "ai_reviewing"
    ] == [(completed, 7) for completed in range(1, 8)]


def test_progress_uses_candidate_issue_count_for_ai_stage(monkeypatch):
    progress = []
    client = AIClient(
        {
            "SH-2": _review(
                "SH-2",
                ("SUMMARY.PROBABILITY", "PASS", "", ""),
            )
        }
    )
    candidate = _issue(
        "SH-2",
        summary="[ACME][T7][V1.1][Video]: Video freezes,often",
    )
    issues = [_issue("SH-1"), candidate, candidate]

    _run(
        monkeypatch,
        issues,
        lambda: client,
        lambda *value: progress.append(value),
    )

    assert len(client.requests) == 1
    assert progress == [
        ("fetching", 3, 3),
        ("rule_auditing", 2, 2),
        ("ai_reviewing", 1, 1),
        ("finalizing", 2, 2),
    ]


def test_large_rule_audit_releases_processed_raw_payloads():
    class HeavyDescription(str):
        pass

    class HeavyIssue(dict):
        pass

    total = 256
    issue_refs = []
    description_refs = []

    def heavy_issues(start, stop):
        for index in range(start, stop):
            description = HeavyDescription(_description())
            issue = HeavyIssue(
                _issue(f"SH-{index + 1}", description=description)
            )
            issue_refs.append(weakref.ref(issue))
            description_refs.append(weakref.ref(description))
            yield issue

    class ReleasingClient:
        page_size = 32

        def search_page(self, _jql, **kwargs):
            start = kwargs.get("start_at", 0)
            stop = min(start + self.page_size, total)
            return SearchPage(
                list(heavy_issues(start, stop)),
                start,
                self.page_size,
                total,
                stop >= total,
            )

    halfway_liveness = []
    post_rule_liveness = []

    def record_liveness(stage, processed, _total):
        current = (
            sum(reference() is not None for reference in issue_refs),
            sum(reference() is not None for reference in description_refs),
        )
        if stage == "fetching" and processed == total // 2:
            halfway_liveness.append(current)
        if stage == "ai_reviewing":
            post_rule_liveness.append(current)

    JiraAuditService(ReleasingClient(), base_url="https://jira.example.com").run(
        ResolvedAuditInput("jql", "project = SH", "project = SH"),
        record_liveness,
    )

    assert halfway_liveness == [(0, 0)]
    assert post_rule_liveness == [(0, 0)]


def test_large_query_filters_before_rules_and_requests_only_required_fields(
    monkeypatch,
):
    class ForbiddenDescription:
        def __str__(self):
            raise AssertionError("rejected descriptions must not be normalized")

    class LargePagedClient:
        total = 6484
        eligible = 177
        page_size = 100

        def __init__(self):
            self.search_calls = []

        def search_page(self, _jql, **kwargs):
            self.search_calls.append(kwargs)
            start = kwargs.get("start_at", 0)
            stop = min(start + self.page_size, self.total)
            issues = [
                _issue(
                    f"SH-{index + 1}",
                    creator={"displayName": (
                        "Chao Li" if index < self.eligible else "Someone Else"
                    )},
                    description=(
                        _description()
                        if index < self.eligible
                        else ForbiddenDescription()
                    ),
                )
                for index in range(start, stop)
            ]
            return SearchPage(
                issues,
                start,
                self.page_size,
                self.total,
                stop >= self.total,
            )

    monkeypatch.setattr(
        service_module,
        "create_chat_client",
        lambda: pytest.fail("passing eligible issues do not require AI"),
    )
    client = LargePagedClient()

    report = JiraAuditService(
        client,
        base_url="https://jira.example.com",
    ).run(
        ResolvedAuditInput(
            "jql",
            "private query marker",
            "private query marker",
        ),
        lambda *_value: None,
    )

    assert report.total_count == 177
    assert [item.key for item in report.issues[:2]] == ["SH-1", "SH-2"]
    assert report.issues[-1].key == "SH-177"
    assert {
        tuple(call["fields"])
        for call in client.search_calls
    } == {
        ("creator", "summary", "description", "reporter", "components")
    }
