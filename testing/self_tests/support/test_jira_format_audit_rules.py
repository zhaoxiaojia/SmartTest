from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from support.jira_integration.audit import (
    AuditReport,
    AuditViolation,
    IssueAuditResult,
    JiraAuditService,
    ResolvedAuditInput,
    active_rules,
    audit_issue,
    export_audit_xlsx,
    resolve_audit_input,
)
from support.jira_integration.audit.rules import (
    QA_CREATOR_NAMES,
    ai_reviewable_violations,
    is_audit_eligible,
)
from support.jira_integration.core.models import SearchPage


RULE_IDS = tuple(
    """
    SUMMARY.FORMAT SUMMARY.PROBABILITY COMPONENT.REQUIRED
    DESCRIPTION.STEPS_TO_REPRODUCE DESCRIPTION.ACTUAL_RESULTS
    DESCRIPTION.EXPECTED_RESULTS DESCRIPTION.COMPARISON DESCRIPTION.NOTES
    DESCRIPTION.RATE_FORMAT DESCRIPTION.NOTES_HW DESCRIPTION.NOTES_SW
    """.split()
)
_DEFAULT_CREATOR = object()


def _description(
    *,
    steps="1. Start playback.\n2. Seek to 00:30.",
    actual="Video freezes.",
    expected="Playback continues.",
    rate="2/2",
    comparison="Previous version V1.0 is normal; current version V1.1 is broken.",
    notes="HW info: T7 reference board\nSW info: V1.1",
):
    values = (
        ("Steps to reproduce", steps),
        ("Actual results", actual),
        ("Expected results", expected),
        ("Reproducibility rate", rate),
        ("Comparison", comparison),
        ("Notes", notes),
    )
    return "\n".join(f"[{heading}]:\n{value}" for heading, value in values)


def _issue(
    key="SH-123",
    *,
    summary="[ACME][T7][V1.1][Video]: Video freezes after seeking,2/2",
    description=None,
    components=("Video",),
    labels=("regression",),
    creator=_DEFAULT_CREATOR,
):
    return {
        "key": key,
        "fields": {
            "summary": summary,
            "description": _description() if description is None else description,
            "reporter": {"displayName": "Coco"},
            "creator": (
                {"displayName": "Chao Li"}
                if creator is _DEFAULT_CREATOR
                else creator
            ),
            "components": [{"name": value} for value in components],
            "labels": list(labels),
        },
    }


def _violations(**changes):
    return {
        item.rule_id
        for item in audit_issue(_issue(**changes), base_url="https://jira.example.com").violations
    }


def test_complete_issue_accepts_four_five_and_six_summary_groups():
    summaries = (
        "[ACME][T7][V1.1][Video]: Video freezes,50%",
        "[SH-1][ACME][T7][V1.1][Video]: Video freezes,1/2",
        "[SH-1][BUG-2][ACME][T7][V1.1][Video]: Video freezes,100%.",
        "【ACME】 [T7] 【V1.1】 [Video]： Video freezes,50%",
    )

    assert tuple(rule.rule_id for rule in active_rules()) == RULE_IDS
    assert all(audit_issue(_issue(summary=value), base_url="https://jira.example.com").passed
               for value in summaries)


EMPTY_DESCRIPTION = _description(
    steps="",
    actual="",
    expected="",
    rate="",
    comparison="",
    notes="",
)


@pytest.mark.parametrize(
    ("changes", "expected"),
    (
        ({"summary": "invalid"}, {"SUMMARY.FORMAT"}),
        (
            {"summary": "[ ][ ][ ][Video]: Video freezes,50%"},
            {"SUMMARY.FORMAT"},
        ),
        ({"summary": "[客户][T7][V1][Video]: 播放冻结,50%"}, set()),
        (
            {"summary": "[ACME][T7][V1][Video]: Video freezes,3/2"},
            {"SUMMARY.PROBABILITY"},
        ),
        ({"components": ()}, {"COMPONENT.REQUIRED"}),
        (
            {"description": EMPTY_DESCRIPTION, "labels": ()},
            {
                "DESCRIPTION.STEPS_TO_REPRODUCE",
                "DESCRIPTION.ACTUAL_RESULTS",
                "DESCRIPTION.EXPECTED_RESULTS",
                "DESCRIPTION.COMPARISON",
                "DESCRIPTION.NOTES",
                "DESCRIPTION.NOTES_HW",
                "DESCRIPTION.NOTES_SW",
            },
        ),
        (
            {
                "description": _description(
                    rate="3/2",
                    notes="HW info:\nSW info:",
                ),
                "labels": (),
            },
            {
                "DESCRIPTION.RATE_FORMAT",
                "DESCRIPTION.NOTES_HW",
                "DESCRIPTION.NOTES_SW",
            },
        ),
        ({"description": _description(comparison="V1 and V2 compared.")}, set()),
    ),
)
def test_rule_failure_matrix(changes, expected):
    assert expected == _violations(**changes)


@pytest.mark.parametrize(
    ("rule_id", "changes", "source_field"),
    (
        ("SUMMARY.FORMAT", {"summary": "invalid summary"}, "summary"),
        (
            "SUMMARY.PROBABILITY",
            {"summary": "[ACME][T7][V1][Video]: Video freezes,often"},
            "summary",
        ),
        ("COMPONENT.REQUIRED", {"components": ()}, None),
        (
            "DESCRIPTION.STEPS_TO_REPRODUCE",
            {"description": _description(steps="")},
            "description",
        ),
        (
            "DESCRIPTION.ACTUAL_RESULTS",
            {"description": _description(actual="")},
            "description",
        ),
        (
            "DESCRIPTION.EXPECTED_RESULTS",
            {"description": _description(expected="")},
            "description",
        ),
        (
            "DESCRIPTION.COMPARISON",
            {"description": _description(comparison="")},
            "description",
        ),
        (
            "DESCRIPTION.NOTES",
            {"description": _description(notes="")},
            "description",
        ),
        (
            "DESCRIPTION.RATE_FORMAT",
            {"description": _description(rate="often")},
            "description",
        ),
        (
            "DESCRIPTION.NOTES_HW",
            {"description": _description(notes="SW info: V1.1")},
            "description",
        ),
        (
            "DESCRIPTION.NOTES_SW",
            {"description": _description(notes="HW info: T7")},
            "description",
        ),
    ),
)
def test_each_violation_keeps_the_full_original_jira_field(
    rule_id, changes, source_field
):
    issue = _issue(**changes)
    result = audit_issue(issue, base_url="https://jira.example.com")
    violation = next(
        item for item in result.violations if item.rule_id == rule_id
    )
    expected = "[]" if source_field is None else issue["fields"][source_field]

    assert violation.observed == expected
    assert violation.observed


def test_legacy_comparision_heading_is_a_hard_missing_section_violation():
    description = _description().replace("[Comparison]", "[Comparision]")

    result = audit_issue(
        _issue(description=description, labels=()),
        base_url="https://jira.example.com",
    )

    assert {item.rule_id for item in result.violations} == {
        "DESCRIPTION.COMPARISON"
    }
    assert not ai_reviewable_violations(result)


def test_only_declared_ambiguous_violations_become_ai_candidates():
    hard_failure = audit_issue(
        _issue(description=_description(actual="")),
        base_url="https://jira.example.com",
    )
    fuzzy_failures = (
        audit_issue(
            _issue(
                summary="[ACME][T7][V1.1][Video]: Video freezes,often",
                labels=(),
            ),
            base_url="https://jira.example.com",
        ),
        audit_issue(
            _issue(description=_description(rate="intermittent"), labels=()),
            base_url="https://jira.example.com",
        ),
        audit_issue(
            _issue(description=_description(comparison=""), labels=()),
            base_url="https://jira.example.com",
        ),
        audit_issue(
            _issue(
                description=_description(
                    notes="Hardware information:\nSoftware information: V1.1"
                ),
                labels=(),
            ),
            base_url="https://jira.example.com",
        ),
        audit_issue(
            _issue(
                description=_description(notes="HW info: T7\nSW info:"),
                labels=(),
            ),
            base_url="https://jira.example.com",
        ),
    )

    assert not ai_reviewable_violations(hard_failure)
    assert [bool(ai_reviewable_violations(result)) for result in fuzzy_failures] == [
        True,
        True,
        True,
        True,
        True,
    ]
    full_description = _description(
        actual="HW info: T7 reference board\nVideo freezes.",
        notes="SW info: V1.1",
    )
    full_context_result = audit_issue(
        _issue(description=full_description, labels=()),
        base_url="https://jira.example.com",
    )
    assert [item.rule_id for item in ai_reviewable_violations(
        full_context_result,
        description=full_description,
    )] == ["DESCRIPTION.NOTES_HW"]


def test_active_rules_and_validation_match_root_effective_behavior():
    import jira_handler

    probes = (
        _issue(),
        _issue(summary="invalid"),
        _issue(summary="[ ][ ][ ][Video]: Video freezes,50%"),
        _issue(summary="[客户][T7][V1][Video]: 播放冻结,50%"),
        _issue(summary="[ACME][T7][V1][Video]: Video freezes,often"),
        _issue(components=()),
        _issue(description=""),
        _issue(description=_description(rate="often")),
        _issue(description=_description(notes="")),
        _issue(description=_description(comparison="V1 and V2 compared.")),
    )
    root_emitted = set()
    mismatches = []
    for issue in probes:
        root_rule_ids = {
            violation["rule_id"]
            for violation in jira_handler.validate_issue(
                issue,
                base_url="https://jira.example.com",
            )["violations"]
        } - {"REGRESSION.EVIDENCE"}
        smarttest_rule_ids = {
            violation.rule_id
            for violation in audit_issue(
                issue,
                base_url="https://jira.example.com",
            ).violations
        }
        root_emitted.update(root_rule_ids)
        if smarttest_rule_ids != root_rule_ids:
            mismatches.append((root_rule_ids, smarttest_rule_ids))

    assert mismatches == []
    assert {rule.rule_id for rule in active_rules()} == root_emitted


def test_active_rule_text_does_not_reintroduce_english_only_semantics():
    rules = {rule.rule_id: rule for rule in active_rules()}
    summary_format = rules["SUMMARY.FORMAT"]

    assert summary_format.requirement == (
        "Summary 必须包含 4–6 个方括号分组，最后四组依次为客户、CHIP、"
        "系统版本和模块，冒号后填写问题描述与复现概率。"
    )
    assert summary_format.guidance == (
        "使用“[客户][CHIP][版本][模块]: 问题描述,复现概率”；"
        "前面可选增加公共 Jira ID 和客户 Bug ID。"
    )
    assert all(
        "英文" not in rule.requirement + rule.guidance
        and "English" not in rule.requirement + rule.guidance
        for rule in rules.values()
    )


@pytest.mark.parametrize(
    "creator",
    (
        {"displayName": "  CHAO   LI  "},
        {"name": r"AMLOGIC\chao.li"},
        {"key": "chao.li@example.com"},
        {"accountId": "chao.li"},
        "Chao Li",
        {"displayName": "Someone Else", "name": "chao.li"},
        {"displayName": "Someone Else"},
        None,
    ),
)
def test_creator_eligibility_matches_portable_root_handler(creator):
    import jira_handler

    issue = _issue(creator=creator)
    normalized = jira_handler.normalize_issue(issue)
    expected = normalized["creator_match_name"] in {
        name.casefold() for name in jira_handler.QA_CREATOR_NAMES
    }

    assert is_audit_eligible(issue) is expected


def test_creator_allowlist_matches_portable_root_handler():
    import jira_handler

    assert QA_CREATOR_NAMES == jira_handler.QA_CREATOR_NAMES


@dataclass
class FakeClient:
    pages: list[SearchPage] | None = None
    filter_payload: dict | None = None
    error: Exception | None = None

    def __post_init__(self):
        self.search_calls = []
        self.filter_calls = []

    def search_page(self, jql, **kwargs):
        self.search_calls.append((jql, kwargs))
        if self.error:
            raise self.error
        if self.pages:
            return self.pages.pop(0)
        return SearchPage([], 0, kwargs.get("max_results", 1), 0, True)

    def fetch_filter(self, filter_id):
        self.filter_calls.append(filter_id)
        return dict(self.filter_payload or {})


@pytest.mark.parametrize(
    ("text", "kind", "jql", "filter_payload"),
    (
        ("project = SH", "jql", "project = SH", None),
        ("https://jira.example.com/browse/SH-123", "issue_url", 'key = "SH-123"', None),
        ("https://jira.example.com/issues/?jql=project%20%3D%20SH",
         "jql_url", "project = SH", None),
        ("https://jira.example.com/filter/42", "filter_url", "labels = regression",
         {"jql": "labels = regression"}),
    ),
)
def test_resolves_supported_jql_and_jira_urls(text, kind, jql, filter_payload):
    client = FakeClient(filter_payload=filter_payload)

    result = resolve_audit_input(text, base_url="https://jira.example.com", client=client)

    assert (result.source_kind, result.jql) == (kind, jql)
    assert client.search_calls[0][0] == jql
    assert client.search_calls[0][1]["fields"] == (
        "creator",
        "summary",
        "description",
        "reporter",
        "components",
    )


def test_rejects_empty_external_or_invalid_jira_input():
    invalid = (
        ("", "JQL"),
        ("https://other.example.com/browse/SH-1", "host"),
        ("ftp://jira.example.com/browse/SH-1", "HTTP"),
        ("https://jira.example.com/browse/not-a-key", "URL"),
    )
    for text, message in invalid:
        with pytest.raises(ValueError, match=message):
            resolve_audit_input(text, base_url="https://jira.example.com", client=FakeClient())
    with pytest.raises(ValueError, match="JQL"):
        resolve_audit_input("project =", base_url="https://jira.example.com",
                            client=FakeClient(error=RuntimeError("invalid")))


def test_service_paginates_and_reports_stable_progress_stages():
    client = FakeClient(
        pages=[
            SearchPage([_issue("SH-1"), _issue("SH-2")], 0, 2, 5, False),
            SearchPage([_issue("SH-2"), _issue("SH-3")], 0, 2, 5, False),
            SearchPage([_issue("SH-4")], 4, 2, 5, True),
        ]
    )
    progress = []

    report = JiraAuditService(client, base_url="https://jira.example.com").run(
        ResolvedAuditInput("jql", "project = SH", "project = SH"),
        lambda *value: progress.append(value),
    )

    assert [call[1]["start_at"] for call in client.search_calls] == [0, 2, 4]
    assert [item.key for item in report.issues] == ["SH-1", "SH-2", "SH-3", "SH-4"]
    assert report.passed_count == 4
    assert progress == [
        ("fetching", 2, 5),
        ("fetching", 4, 5),
        ("fetching", 5, 5),
        ("rule_auditing", 4, 4),
        ("ai_reviewing", 0, 0),
        ("finalizing", 4, 4),
    ]


def test_empty_service_run_reports_every_stable_progress_stage():
    progress = []

    JiraAuditService(FakeClient(), base_url="https://jira.example.com").run(
        ResolvedAuditInput("jql", "project = NONE", "project = NONE"),
        lambda *value: progress.append(value),
    )

    assert progress == [
        ("fetching", 0, 0),
        ("rule_auditing", 0, 0),
        ("ai_reviewing", 0, 0),
        ("finalizing", 0, 0),
    ]


def test_openpyxl_export_matches_root_reader_visible_report_contract(tmp_path):
    import jira_handler

    summary = "[ACME][T7][V1.1][Video]: Video freezes,often"
    dynamic_reason = "Summary probability is invalid."
    dynamic_guidance = "Use a percentage or fraction."
    rules_by_id = {rule.rule_id: rule for rule in active_rules()}
    format_rule = rules_by_id["SUMMARY.FORMAT"]
    format_violation = AuditViolation(
        format_rule.rule_id,
        format_rule.section,
        format_rule.field,
        summary,
        "Summary grouping is invalid.",
        format_rule.guidance,
    )
    component_violation = AuditViolation(
        "COMPONENT.REQUIRED",
        "Component",
        "Component",
        "[]",
        "This supplied reason must be replaced by report text.",
        "This supplied guidance must be replaced by report text.",
    )
    probability_violation = AuditViolation(
        "SUMMARY.PROBABILITY",
        "Summary",
        "Summary.Probability",
        summary,
        dynamic_reason,
        dynamic_guidance,
    )
    issues = (
        IssueAuditResult(
            "SH-2",
            "https://jira.example.com/browse/SH-2",
            summary,
            "Bob",
            False,
            (component_violation, probability_violation, format_violation),
        ),
        IssueAuditResult(
            "SH-1",
            "https://jira.example.com/browse/SH-1",
            "Valid summary",
            "Bob",
            False,
            (component_violation,),
        ),
        IssueAuditResult(
            "SH-3",
            "https://jira.example.com/browse/SH-3",
            "Valid summary",
            "Alice",
            True,
            (),
        ),
    )
    root_rows = [
        {
            "key": issue.key,
            "url": issue.url,
            "reporter": issue.reporter,
            "overall_result": "PASS" if issue.passed else "FAIL",
            "violations": [
                {
                    "rule_id": violation.rule_id,
                    "spec_section": violation.section,
                    "jira_field": {
                        "COMPONENT.REQUIRED": "Component",
                        "SUMMARY.PROBABILITY": "Summary.Probability",
                        "SUMMARY.FORMAT": "Summary",
                    }[violation.rule_id],
                    "jira_value": violation.observed,
                    "failure_reason": violation.reason,
                    "correction_guidance": violation.guidance,
                }
                for violation in issue.violations
            ],
        }
        for issue in issues
    ]
    generated_at = datetime(2026, 7, 24, 10, 11, 12, tzinfo=timezone.utc)
    jql = "=project = SH"
    report = AuditReport(
        resolved=ResolvedAuditInput("jql", jql, jql),
        generated_at=generated_at,
        rules=active_rules(),
        issues=issues,
    )
    root_path = jira_handler.export_xlsx(
        root_rows,
        tmp_path / "root.xlsx",
        jql=jql,
        generated_at=generated_at.isoformat(),
    )
    smart_dir = tmp_path / "smart"
    first = export_audit_xlsx(report, downloads_dir=smart_dir, now=generated_at)
    second = export_audit_xlsx(report, downloads_dir=smart_dir, now=generated_at)
    root_workbook = load_workbook(root_path)
    smart_workbook = load_workbook(first)
    root_detail = root_workbook["违规明细"]
    format_row = next(
        row for row in range(2, root_detail.max_row + 1)
        if root_detail.cell(row, 4).value == "SUMMARY.FORMAT"
    )
    assert "英文问题描述" in root_detail.cell(format_row, 6).value
    root_detail.cell(format_row, 6).value = format_rule.requirement

    def values(sheet):
        return tuple(
            tuple(None if cell.value == "" else cell.value for cell in row)
            for row in sheet.iter_rows()
        )

    assert smart_workbook.sheetnames == root_workbook.sheetnames
    for sheet_name in root_workbook.sheetnames:
        smart_sheet = smart_workbook[sheet_name]
        root_sheet = root_workbook[sheet_name]
        assert values(smart_sheet) == values(root_sheet)
        assert smart_sheet.freeze_panes == root_sheet.freeze_panes == "A2"
        assert smart_sheet.auto_filter.ref == root_sheet.auto_filter.ref
        assert {
            str(value) for value in smart_sheet.merged_cells.ranges
        } == {str(value) for value in root_sheet.merged_cells.ranges}
        columns = "ABC" if sheet_name == "汇总" else "ABCDEFGHIJ"
        assert tuple(
            smart_sheet.column_dimensions[column].width for column in columns
        ) == tuple(
            root_sheet.column_dimensions[column].width for column in columns
        )
        for address in ("A1", "A2"):
            smart_cell, root_cell = smart_sheet[address], root_sheet[address]
            assert smart_cell.font.name == root_cell.font.name == "Calibri"
            assert smart_cell.font.sz == root_cell.font.sz == 11
            assert smart_cell.font.bold == root_cell.font.bold
            assert smart_cell.fill.fgColor.rgb == root_cell.fill.fgColor.rgb
            assert smart_cell.alignment.horizontal == root_cell.alignment.horizontal
            assert smart_cell.alignment.vertical == root_cell.alignment.vertical
            assert smart_cell.alignment.wrap_text == root_cell.alignment.wrap_text
    assert first.name == "jira_format_audit_20260724_101112.xlsx"
    assert second.name == "jira_format_audit_20260724_101112_2.xlsx"
    assert not list(smart_dir.glob("*.tmp"))
    assert smart_workbook["汇总"]["B3"].value == jql
    assert smart_workbook["汇总"]["B3"].data_type == "s"
    detail_text = " ".join(
        str(cell.value or "")
        for row in smart_workbook["违规明细"].iter_rows()
        for cell in row
    )
    assert "REGRESSION.EVIDENCE" not in detail_text
    assert "英文问题描述" not in detail_text
