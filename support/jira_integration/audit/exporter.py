from __future__ import annotations

import ctypes
import os
import tempfile
from ctypes import wintypes
from datetime import datetime
from pathlib import Path
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .models import AuditReport


_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def export_audit_xlsx(
    report: AuditReport,
    *,
    downloads_dir: Path | None = None,
    now: datetime | None = None,
) -> Path:
    target_dir = (
        Path(downloads_dir) if downloads_dir is not None else _windows_downloads_dir()
    )
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    target = _unique_path(
        target_dir,
        f"jira_format_audit_{timestamp}",
        ".xlsx",
    )
    fd, temporary_name = tempfile.mkstemp(
        prefix=".jira_format_audit_",
        suffix=".tmp",
        dir=target_dir,
    )
    os.close(fd)
    temporary = Path(temporary_name)
    try:
        _build_workbook(report).save(temporary)
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()
    return target.resolve()


def _build_workbook(report: AuditReport) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _append_rows(
        summary,
        [
            ["Metric", "Value"],
            ["Generated at", report.generated_at.isoformat(sep=" ", timespec="seconds")],
            ["Source", report.resolved.source_kind],
            ["Original input", report.resolved.original],
            ["JQL", report.resolved.jql],
            ["Total issues", report.total_count],
            ["Passed issues", report.passed_count],
            ["Failed issues", report.failed_count],
            ["Violations", report.violation_count],
        ],
    )

    rules = workbook.create_sheet("Rules")
    _append_rows(
        rules,
        [["Rule ID", "Section", "Field", "Requirement", "Guidance"]]
        + [
            [
                rule.rule_id,
                rule.section,
                rule.field,
                rule.requirement,
                rule.guidance,
            ]
            for rule in report.rules
        ],
    )

    issues = workbook.create_sheet("Issues")
    _append_rows(
        issues,
        [["Key", "URL", "Summary", "Reporter", "Status", "Violation count"]]
        + [
            [
                issue.key,
                issue.url,
                issue.summary,
                issue.reporter,
                "PASS" if issue.passed else "FAIL",
                len(issue.violations),
            ]
            for issue in report.issues
        ],
    )
    for row in range(2, issues.max_row + 1):
        _set_hyperlink(issues.cell(row=row, column=2))

    violations = workbook.create_sheet("Violations")
    violation_rows = [
        [
            "Key",
            "URL",
            "Rule ID",
            "Section",
            "Field",
            "Observed",
            "Requirement",
            "Reason",
            "Guidance",
        ]
    ]
    rule_by_id = {rule.rule_id: rule for rule in report.rules}
    for issue in report.issues:
        for violation in issue.violations:
            rule = rule_by_id.get(violation.rule_id)
            violation_rows.append(
                [
                    issue.key,
                    issue.url,
                    violation.rule_id,
                    violation.section,
                    violation.field,
                    violation.observed,
                    rule.requirement if rule else "",
                    violation.reason,
                    violation.guidance,
                ]
            )
    _append_rows(violations, violation_rows)
    for row in range(2, violations.max_row + 1):
        _set_hyperlink(violations.cell(row=row, column=2))

    for sheet in workbook.worksheets:
        _style_sheet(sheet)
    return workbook


def _append_rows(sheet: Worksheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)


def _set_hyperlink(cell) -> None:
    if cell.value:
        cell.hyperlink = str(cell.value)
        cell.style = "Hyperlink"


def _style_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 48)
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _unique_path(directory: Path, stem: str, suffix: str) -> Path:
    candidate = directory / f"{stem}{suffix}"
    counter = 2
    while candidate.exists():
        candidate = directory / f"{stem}_{counter}{suffix}"
        counter += 1
    return candidate


def _windows_downloads_dir() -> Path:
    if os.name != "nt":
        return Path.home() / "Downloads"

    class GUID(ctypes.Structure):
        _fields_ = [
            ("Data1", wintypes.DWORD),
            ("Data2", wintypes.WORD),
            ("Data3", wintypes.WORD),
            ("Data4", ctypes.c_ubyte * 8),
        ]

        @classmethod
        def from_uuid(cls, value: UUID):
            raw = value.bytes_le
            return cls(
                int.from_bytes(raw[0:4], "little"),
                int.from_bytes(raw[4:6], "little"),
                int.from_bytes(raw[6:8], "little"),
                (ctypes.c_ubyte * 8)(*raw[8:]),
            )

    folder_id = GUID.from_uuid(UUID("374DE290-123F-4565-9164-39C4925E467B"))
    result_path = ctypes.c_wchar_p()
    status = ctypes.windll.shell32.SHGetKnownFolderPath(  # type: ignore[attr-defined]
        ctypes.byref(folder_id),
        0,
        None,
        ctypes.byref(result_path),
    )
    if status != 0 or not result_path.value:
        raise OSError(f"无法获取 Windows Downloads Known Folder：{status}")
    try:
        return Path(result_path.value)
    finally:
        ctypes.windll.ole32.CoTaskMemFree(result_path)  # type: ignore[attr-defined]
