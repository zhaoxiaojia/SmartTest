from __future__ import annotations

import os
from pathlib import Path
import tempfile

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
_STATUS_FILLS = {
    "已更新": PatternFill("solid", fgColor="FFC6EFCE"),
    "未更新": PatternFill("solid", fgColor="FFFFEB9C"),
    "格式有误": PatternFill("solid", fgColor="FFFFC7CE"),
}


def write_xlsx_table(
    output_path,
    *,
    sheet_name,
    headers,
    rows,
    hyperlinks=None,
):
    """Write one reusable, styled XLSX table through the global report owner."""
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = str(sheet_name)
    sheet.append([_clean(value) for value in headers])
    for row in rows:
        sheet.append([_clean(value) for value in row])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            fill = _STATUS_FILLS.get(cell.value)
            if fill is not None:
                cell.fill = fill
    for (row, column), url in (hyperlinks or {}).items():
        cell = sheet.cell(row=row, column=column)
        cell.hyperlink = str(url)
        cell.style = "Hyperlink"
    for index, column in enumerate(sheet.columns, 1):
        width = max(
            (len(str(cell.value or "")) for cell in column),
            default=0,
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(
            45, max(12, width + 2),
        )

    _save_workbook(workbook, target)
    return target.resolve()


def write_xlsx_sections(output_path, *, sheet_name, sections):
    """Write repeated group/header/data sections with one global XLSX owner."""
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = str(sheet_name)
    hyperlinks = {}
    header_rows = []
    group_rows = []
    for section in sections:
        sheet.append([_clean(value) for value in section["group"]])
        group_rows.append(sheet.max_row)
        sheet.append([_clean(value) for value in section["headers"]])
        header_rows.append(sheet.max_row)
        for row_index, row in enumerate(section["rows"]):
            sheet.append([_clean(value) for value in row])
            for column, url in section.get("hyperlinks", {}).get(row_index, {}).items():
                hyperlinks[(sheet.max_row, column)] = url

    sheet.freeze_panes = "A3"
    for row_index in group_rows:
        for cell in sheet[row_index]:
            cell.fill = PatternFill("solid", fgColor="FFD9EAF7")
            cell.font = Font(name="Calibri", size=11, bold=True, color="FF1F1F1F")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_index in header_rows:
        for cell in sheet[row_index]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
            )
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row not in header_rows and cell.row not in group_rows:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                fill = _STATUS_FILLS.get(cell.value)
                if fill is not None:
                    cell.fill = fill
    for (row, column), url in hyperlinks.items():
        cell = sheet.cell(row=row, column=column)
        cell.hyperlink = str(url)
        cell.style = "Hyperlink"
    for index, column in enumerate(sheet.columns, 1):
        width = max((len(str(cell.value or "")) for cell in column), default=0)
        sheet.column_dimensions[get_column_letter(index)].width = min(
            55, max(12, width + 2),
        )
    _save_workbook(workbook, target)
    return target.resolve()


def _save_workbook(workbook, target):
    handle, temporary_name = tempfile.mkstemp(
        prefix=f".{target.stem}-", suffix=".tmp", dir=target.parent,
    )
    os.close(handle)
    temporary = Path(temporary_name)
    try:
        workbook.save(temporary)
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()


def _clean(value):
    return ILLEGAL_CHARACTERS_RE.sub("", value) if isinstance(value, str) else value
