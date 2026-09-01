from __future__ import annotations

from dataclasses import replace
from datetime import datetime, timezone

from openpyxl import load_workbook
import pytest

from core.domain.detail import DetailSection
from core.domain.values import NamedValue, PersonRef
from core.jira.audit import (
    JiraAuditUseCase,
    audit_issue,
    export_audit_xlsx,
    resolve_audit_input,
)
from core.jira.domain import Issue, IssueIdentity, JiraProjectRef, RichText


GOOD_DESCRIPTION = """Steps to reproduce: Open video
Actual results: Freeze
Expected results: Playback
Reproducibility rate: 1/2
Comparison: Previous build works
Notes:
HW info: board A
SW info: build 1
"""


def _issue(key: str, creator: str, description: str = GOOD_DESCRIPTION) -> Issue:
    return Issue(
        IssueIdentity(key, key, f"https://jira.example/browse/{key}"),
        "[ACME][T7][V1][Video]: freezes", JiraProjectRef("SH"),
        NamedValue("1", "Open"), NamedValue("2", "Bug"),
        creator=PersonRef(creator.casefold(), creator.casefold(), creator),
        components=(NamedValue("10", "Video"),),
        description=DetailSection.loaded(RichText(description)),
    )


def test_jira_input_resolves_issue_search_and_filter_urls_and_validates_jql() -> None:
    validated = []
    fetch_filter = lambda filter_id: {"jql": f"filter = {filter_id}"}
    validate = lambda jql: validated.append(jql)

    issue = resolve_audit_input(
        "https://jira.example/browse/SH-10", base_url="https://jira.example",
        fetch_filter=fetch_filter, validate_jql=validate,
    )
    search = resolve_audit_input(
        "https://jira.example/issues/?jql=project%20%3D%20SH",
        base_url="https://jira.example", fetch_filter=fetch_filter,
        validate_jql=validate,
    )
    filtered = resolve_audit_input(
        "https://jira.example/filter/42", base_url="https://jira.example",
        fetch_filter=fetch_filter, validate_jql=validate,
    )
    filter_id = resolve_audit_input(
        "43", base_url="https://jira.example",
        fetch_filter=fetch_filter, validate_jql=validate,
    )

    assert (issue.source_kind, issue.jql) == ("issue_url", 'key = "SH-10"')
    assert (search.source_kind, search.jql) == ("jql_url", "project = SH")
    assert (filtered.source_kind, filtered.jql) == ("filter_url", "filter = 42")
    assert (filter_id.source_kind, filter_id.jql) == ("filter_id", "filter = 43")
    assert validated == ['key = "SH-10"', "project = SH", "filter = 42", "filter = 43"]
    with pytest.raises(ValueError, match="host"):
        resolve_audit_input(
            "https://other.example/browse/SH-1", base_url="https://jira.example",
            fetch_filter=fetch_filter, validate_jql=validate,
        )


def test_jira_deterministic_rules_keep_old_table_and_standard_behaviour() -> None:
    assert audit_issue(_issue("SH-1", "Chao Li")).passed
    table = "\n".join((
        "||||模块|||需要填写信息||||测试信息||||",
        "|平台信息|客户/项目代号|T6X高刷|", "|测试环境|测试仪器|-|",
    ))
    assert audit_issue(_issue("SH-2", "Chao Li", table)).passed

    failed = audit_issue(replace(
        _issue("SH-3", "Chao Li", ""),
        summary="bad", components=(),
    ))
    assert {item.rule_id for item in failed.violations} >= {
        "SUMMARY.FORMAT", "COMPONENT.REQUIRED",
        "DESCRIPTION.STEPS_TO_REPRODUCE", "DESCRIPTION.NOTES_HW",
    }


def test_jira_use_case_loads_description_only_for_eligible_creator() -> None:
    eligible = replace(_issue("SH-1", "Chao Li"), description=DetailSection())
    ignored = replace(_issue("SH-2", "Outside User"), description=DetailSection())

    class Source:
        def __init__(self): self.loaded = []
        def list_issues(self, _scope, _cancellation): return eligible, ignored
        def load_details(self, issue, details):
            self.loaded.append((issue.identity.key, details.sections()))
            return replace(issue, description=DetailSection.loaded(RichText(GOOD_DESCRIPTION)))

    source = Source()
    report = JiraAuditUseCase(source).run(
        type("Scope", (), {"source_kind": "jql", "original": "x", "jql": "x"})(),
    )

    assert [item.key for item in report.issues] == ["SH-1"]
    assert source.loaded == [("SH-1", ("description",))]


def test_jira_exporter_creates_summary_and_violation_workbook(tmp_path) -> None:
    issue = replace(_issue("SH-1", "Chao Li", ""), components=())

    class Source:
        def list_issues(self, _scope, _cancellation): return (issue,)
        def load_details(self, issue, _details): return issue

    report = JiraAuditUseCase(Source()).run(
        type("Scope", (), {"source_kind": "jql", "original": "x", "jql": "project=SH"})(),
    )
    path = export_audit_xlsx(
        report, output_path=tmp_path / "jira.xlsx",
        now=datetime(2026, 8, 29, tzinfo=timezone.utc),
    )
    workbook = load_workbook(path)

    assert workbook.sheetnames == ["汇总", "违规明细"]
    assert tuple(workbook["汇总"].values)[8] == ("创建人", "违规 Jira 数量", "违规 Jira 号")


def test_jira_use_case_starts_every_detail_when_requests_are_below_capacity() -> None:
    from threading import Barrier
    from core.async_tasks import AsyncTaskManager

    first = replace(_issue("SH-1", "Chao Li"), description=DetailSection())
    second = replace(_issue("SH-2", "Chao Li"), description=DetailSection())
    barrier = Barrier(2, timeout=1)

    class Source:
        def list_issues(self, _scope, _cancellation): return first, second
        def load_details(self, issue, _details):
            barrier.wait()
            return replace(issue, description=DetailSection.loaded(RichText(GOOD_DESCRIPTION)))

    manager = AsyncTaskManager(max_workers=3)
    try:
        root = manager.submit_coordinator(
            "jira-review", lambda token, _progress: JiraAuditUseCase(Source()).run(
                type("Scope", (), {"source_kind": "jql", "original": "x", "jql": "x"})(),
                task_manager=manager, parent_task_id=token.task_id,
            ),
        )
        report = root.result(timeout=2)
        root_id = manager.task_id(root)
        assert [item.key for item in report.issues] == ["SH-1", "SH-2"]
        assert len([task for task in manager._tasks.values() if task.parent_id == root_id]) == 2
    finally:
        manager.close()


def test_jira_root_cancellation_cancels_scheduled_issue_children() -> None:
    from core.async_tasks import AsyncTaskManager, TaskCancelled

    issues = tuple(replace(_issue(f"SH-{index}", "Chao Li"), description=DetailSection()) for index in (1, 2))
    manager = AsyncTaskManager(max_workers=3)
    root = manager.register_long_running("jira-review")

    class Source:
        def list_issues(self, _scope, _cancellation): return issues
        def load_details(self, issue, _details):
            if issue.identity.key == "SH-1":
                manager.cancel(root)
            return replace(issue, description=DetailSection.loaded(RichText(GOOD_DESCRIPTION)))

    try:
        with pytest.raises(TaskCancelled):
            JiraAuditUseCase(Source()).run(
                type("Scope", (), {"source_kind": "jql", "original": "x", "jql": "x"})(),
                task_manager=manager, parent_task_id=root,
            )
        assert all(task.token._event.is_set() for task in manager._tasks.values() if task.root_id == root)
    finally:
        manager.close()


def test_jira_review_uses_all_three_available_detail_workers() -> None:
    from threading import Barrier
    from core.async_tasks import AsyncTaskManager

    issues = tuple(replace(_issue(f"SH-{index}", "Chao Li"), description=DetailSection()) for index in (1, 2, 3))
    barrier = Barrier(3, timeout=1)

    class Source:
        def list_issues(self, _scope, _cancellation): return issues
        def load_details(self, issue, _details):
            barrier.wait()
            return replace(issue, description=DetailSection.loaded(RichText(GOOD_DESCRIPTION)))

    manager = AsyncTaskManager(max_workers=3)
    try:
        root = manager.submit_coordinator(
            "jira-review", lambda token, _progress: JiraAuditUseCase(Source()).run(
                type("Scope", (), {"source_kind": "jql", "original": "x", "jql": "x"})(),
                task_manager=manager, parent_task_id=token.task_id,
            ),
        )
        report = root.result(timeout=2)
        assert [item.key for item in report.issues] == ["SH-1", "SH-2", "SH-3"]
    finally:
        manager.close()
