import struct
import subprocess
import sys
import warnings
from pathlib import Path

import matplotlib
from matplotlib import font_manager
from matplotlib import pyplot as plt
from openpyxl import load_workbook
import pytest

ROOT = Path(__file__).resolve().parents[4]

from core.reporting.excel import clean_excel_value, write_excel_workbook
from core.reporting.html import generate_html_report as generate_html_report_from_package
from core.reporting.image import DEFAULT_LINE_CHART_STYLE, LineSeries, render_line_chart
from core.reporting.json import ReportStore
from core.reporting.pdf import export_pdf_report as export_pdf_report_from_package
from core.reporting import (
    build_run_report,
    duration_text,
    export_pdf_report,
    generate_html_report,
    list_reports,
    load_report,
    render_html_report,
    report_file_stem,
    report_html_path,
    report_html_url,
    report_json_path,
    report_pdf_path,
    save_run_report,
    write_xlsx_table,
)
from core.reporting.image.line import _cjk_font_family


def test_report_package_preserves_public_run_report_api():
    assert build_run_report(run_id="r1")["run_id"] == "r1"
    assert duration_text(61_000) == "1m 1s"
    assert all(callable(value) for value in (
        export_pdf_report, generate_html_report, list_reports, load_report,
        render_html_report, report_file_stem, report_html_path, report_html_url,
        report_json_path, report_pdf_path, save_run_report,
    ))
    assert Path(__import__("core.reporting", fromlist=["x"]).__path__[0]).name == "reporting"


def test_global_xlsx_owner_writes_filtered_frozen_wrapped_table(tmp_path):
    output = write_xlsx_table(
        tmp_path / "matrix.xlsx",
        sheet_name="Matrix",
        headers=("Name", "State", "URL"),
        rows=(("Project", "已更新", "https://c/project"),),
        hyperlinks={(2, 3): "https://c/project"},
    )
    sheet = load_workbook(output).active
    assert sheet.title == "Matrix"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:C2"
    assert sheet["A1"].alignment.wrap_text is True
    assert sheet["C2"].hyperlink.target == "https://c/project"


def test_format_packages_expose_public_behaviors(tmp_path):
    assert ReportStore(tmp_path).reports_dir == tmp_path
    assert generate_html_report_from_package is generate_html_report
    assert export_pdf_report_from_package is export_pdf_report


def test_excel_workbook_driver_populates_cleans_and_saves_real_file(tmp_path):
    def populate(workbook):
        summary = workbook.active
        summary.title = "Summary"
        summary["A1"] = clean_excel_value("bad\x00value")
        workbook.create_sheet("Details")

    output = write_excel_workbook(tmp_path / "audit.xlsx", populate)

    workbook = load_workbook(output)
    assert output == (tmp_path / "audit.xlsx").resolve()
    assert workbook.sheetnames == ["Summary", "Details"]
    assert workbook["Summary"]["A1"].value == "badvalue"


def test_report_html_url_regenerates_missing_html_from_saved_json(tmp_path):
    report = build_run_report(run_id="r1", status="passed")
    save_run_report(report, reports_dir=tmp_path)
    html_path = report_html_path("r1", reports_dir=tmp_path)
    html_path.unlink()

    url = report_html_url("r1", reports_dir=tmp_path)

    assert html_path.exists()
    assert url == html_path.resolve().as_uri()


def test_line_chart_writes_real_png_with_configured_dimensions(tmp_path):
    output = render_line_chart(
        ("Jan", "Feb", "Mar"),
        (
            LineSeries("Actual", (10, 14, 12), fill=True),
            LineSeries("Target", (11, 11, 11), color="#D97706"),
        ),
        tmp_path / "trend.png",
        title="Monthly quality",
        highlight_series="Actual",
        kpi_label="Pass rate",
    )

    data = output.read_bytes()
    width, height = struct.unpack(">II", data[16:24])
    assert output == (tmp_path / "trend.png").resolve()
    assert data.startswith(b"\x89PNG\r\n\x1a\n")
    assert (width, height) == (2420, 1100)
    assert len(data) > 1_000


def test_line_chart_prefers_installed_windows_cjk_font_when_available():
    yahei = Path("C:/Windows/Fonts/msyh.ttc")
    if not yahei.is_file():
        pytest.skip("Microsoft YaHei is not installed on this platform")
    family = _cjk_font_family()
    resolved = Path(font_manager.findfont(family, fallback_to_default=False))
    assert family
    assert resolved.samefile(yahei)


def test_line_chart_renders_chinese_without_missing_glyph_warnings(tmp_path):
    if not _cjk_font_family():
        pytest.skip("No supported CJK font is installed")
    with warnings.catch_warnings(record=True) as captured:
        warnings.simplefilter("always")
        output = render_line_chart(
            ("08-02", "08-03"),
            (LineSeries("未关闭 Issue", (120, 126), fill=True),),
            tmp_path / "中文趋势.png",
            title="每日未关闭 Issue 趋势",
            highlight_series="未关闭 Issue",
            kpi_label="当前未关闭",
        )
    assert output.is_file()
    assert not any("Glyph" in str(item.message) and "missing" in str(item.message) for item in captured)


def test_line_chart_does_not_enter_pyplot_gui_path(monkeypatch, tmp_path):
    def fail_gui_path(*_args, **_kwargs):
        raise AssertionError("pyplot GUI path must not be used by background reports")

    monkeypatch.setattr(plt, "subplots", fail_gui_path)

    output = render_line_chart(
        ("Mon", "Tue"),
        (LineSeries("Open", (3, 4)),),
        tmp_path / "background.png",
    )

    assert output.is_file()


def test_default_line_chart_style_preserves_visual_language():
    style = DEFAULT_LINE_CHART_STYLE

    assert style.palette[:3] == ("#4F8EF7", "#2CB67D", "#F59E0B")
    assert style.figure_size == (11.0, 5.0)
    assert (style.figure_dpi, style.save_dpi) == (220, 220)
    assert style.grid_alpha == 0.15
    assert (style.line_width, style.highlight_line_width) == (2.8, 3.0)
    assert style.x_rotation == 20.0


@pytest.mark.parametrize(
    ("labels", "series", "highlight", "message"),
    (
        ((), (LineSeries("Actual", ()),), None, "labels"),
        (("Jan",), (), None, "series"),
        (("Jan", "Feb"), (LineSeries("Actual", (1,)),), None, "length"),
        (("Jan",), (LineSeries("Actual", (1,)),), "Missing", "highlight"),
        (("Jan",), (LineSeries("Actual", (1,)),), None, "kpi.*highlight"),
    ),
)
def test_line_chart_rejects_invalid_input(labels, series, highlight, message, tmp_path):
    with pytest.raises(ValueError, match=message):
        render_line_chart(
            labels,
            series,
            tmp_path / "invalid.png",
            highlight_series=highlight,
            kpi_label="Rate" if message == "kpi.*highlight" else None,
        )


def test_line_chart_does_not_pollute_matplotlib_rcparams(tmp_path):
    keys = ("font.family", "figure.facecolor", "axes.facecolor", "axes.grid")
    before = {key: matplotlib.rcParams[key] for key in keys}
    open_figures = plt.get_fignums()

    render_line_chart(
        ("Jan", "Feb"),
        (LineSeries("Actual", (1, 2)),),
        tmp_path / "isolated.png",
    )

    assert {key: matplotlib.rcParams[key] for key in keys} == before
    assert plt.get_fignums() == open_figures


def test_line_chart_places_title_kpi_and_highlight_at_chart_top(monkeypatch, tmp_path):
    captured = {}

    def capture_figure(figure, *_args, **_kwargs):
        axes = figure.axes[0]
        captured["title"] = axes.title
        captured["texts"] = tuple(axes.texts)
        captured["lines"] = tuple(axes.lines)

    monkeypatch.setattr("matplotlib.figure.Figure.savefig", capture_figure)

    render_line_chart(
        ("Jan", "Feb", "Mar"),
        (
            LineSeries("Actual", (10, 14, 12)),
            LineSeries("Target", (11, 11, 11)),
        ),
        tmp_path / "captured.png",
        title="Monthly quality",
        highlight_series="Actual",
        kpi_label="Pass rate",
    )

    assert captured["title"].get_horizontalalignment() == "left"
    assert captured["title"].get_position()[0] == 0
    assert any(
        text.get_text() == "Pass rate\n12" and text.get_horizontalalignment() == "right"
        for text in captured["texts"]
    )
    highlight_points = [
        line
        for line in captured["lines"]
        if tuple(line.get_xdata()) == (2,)
        and tuple(line.get_ydata()) == (12,)
        and line.get_markeredgecolor() == "white"
    ]
    assert len(highlight_points) == 1
    assert highlight_points[0].get_markersize() > 5

def test_excel_import_does_not_require_or_load_matplotlib_in_fresh_process():
    probe = r'''
import importlib.abc, sys
class BlockMatplotlib(importlib.abc.MetaPathFinder):
    def find_spec(self, fullname, path=None, target=None):
        if fullname == "matplotlib" or fullname.startswith("matplotlib."):
            raise ModuleNotFoundError(fullname)
        return None
sys.meta_path.insert(0, BlockMatplotlib())
from core.reporting.excel import write_excel_workbook
assert callable(write_excel_workbook)
assert not any(name == "matplotlib" or name.startswith("matplotlib.") for name in sys.modules)
print("excel-without-matplotlib")
'''
    result = subprocess.run(
        [sys.executable, "-c", probe], cwd=ROOT,
        capture_output=True, text=True, timeout=15,
    )
    assert result.returncode == 0, result.stderr + result.stdout
    assert "excel-without-matplotlib" in result.stdout


def test_report_root_keeps_lazy_image_public_api():
    from core.reporting import (
        DEFAULT_LINE_CHART_STYLE as root_style,
        LineSeries as root_series,
        render_line_chart as root_renderer,
    )

    assert root_style is DEFAULT_LINE_CHART_STYLE
    assert root_series is LineSeries
    assert root_renderer is render_line_chart
