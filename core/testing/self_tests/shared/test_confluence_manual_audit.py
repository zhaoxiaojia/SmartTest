from __future__ import annotations

from datetime import date, datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
import pytest

from core.confluence.audit import (
    AuditPeriod,
    AuditStatus,
    ConfluencePageDocument,
    ConfluenceWeeklyAuditUseCase,
    UPDATE_MATRIX_POINTS,
    export_audit_xlsx_by_product_line,
    previous_business_week,
)
from core.confluence.audit.models import AuditBatch, AuditFinding, ProjectAudit
from core.confluence.audit.regions import extract_page_region
from core.confluence.project import (
    ConfluencePageRef,
    ProductSpaceRef,
    Project,
    ProjectIdentity,
    ProjectRole,
    SourceEvidence,
)
from core.domain.detail import DetailSection
from core.domain.values import FieldBag, NamedValue, PersonRef


TZ = ZoneInfo("Asia/Shanghai")


def _project(
    project_id="P1", line="DOPL", *, support_mode="", project_status="",
    commercial_date="",
):
    sources = {
        "test_information": "10", "test_plan": "11", "environment": "12",
        "experience": "13", "report_store": "14", "basic": "15",
    }
    return Project(
        ProjectIdentity(project_id, project_id), f"Project {project_id}",
        ProductSpaceRef(line, line), ConfluencePageRef("1", "Catalog", f"https://c/{project_id}"),
        status=NamedValue(name=project_status) if project_status else None,
        support_mode=NamedValue(name=support_mode) if support_mode else None,
        roles=DetailSection.loaded((ProjectRole(
            NamedValue("fae", "FAE QA"),
            (PersonRef("alice", "alice", "Alice"), PersonRef("alice", "alice", "Alice")),
        ),)),
        evidence=DetailSection.loaded(tuple(
            SourceEvidence(name, ConfluencePageRef(page_id, name, f"https://c/{page_id}"))
            for name, page_id in sources.items()
        )),
        facts=DetailSection.loaded(FieldBag.from_mapping({
            "date of commercial approval": commercial_date,
        })),
    )


def _document(page_id, version, when, body):
    return ConfluencePageDocument(page_id, page_id, f"https://c/{page_id}", body, body, version, when)


def test_previous_business_week_and_manual_period_use_exclusive_end() -> None:
    period = previous_business_week(datetime(2026, 8, 29, 12, tzinfo=TZ))

    assert (period.start.date(), period.end.date()) == (date(2026, 8, 17), date(2026, 8, 24))
    assert period.contains(datetime(2026, 8, 23, 23, 59, tzinfo=TZ))
    assert not period.contains(datetime(2026, 8, 24, 0, 0, tzinfo=TZ))


def test_confluence_eight_points_detect_period_updates_and_roles_are_deduplicated() -> None:
    project = _project()
    period = AuditPeriod(
        datetime(2026, 8, 17, tzinfo=TZ), datetime(2026, 8, 24, tzinfo=TZ),
    )

    class Source:
        def refresh_projects(self, _scope): return (project,)
        def load_project_details(self, _project_id, _details, _cancellation): return project
        def load_current_page(self, page_id, _cancellation):
            return _document(page_id, 2, datetime(2026, 8, 20, tzinfo=TZ), _body(page_id, "new"))
        def load_page_versions(self, page_id, _period, _current, _cancellation):
            return (
                _document(page_id, 1, datetime(2026, 8, 10, tzinfo=TZ), _body(page_id, "old")),
                _document(page_id, 2, datetime(2026, 8, 20, tzinfo=TZ), _body(page_id, "new")),
            )

    batch = ConfluenceWeeklyAuditUseCase(Source()).run((project,), period)

    assert len(batch.projects[0].findings) == 8
    assert {finding.status for finding in batch.projects[0].findings} == {AuditStatus.UPDATED}
    assert batch.projects[0].owners == ("Alice",)


def test_confluence_project_failure_does_not_stop_remaining_projects() -> None:
    failed, good = _project("P1"), _project("P2")
    period = AuditPeriod(datetime(2026, 8, 17, tzinfo=TZ), datetime(2026, 8, 24, tzinfo=TZ))

    class Source:
        def load_project_details(self, project_id, _details, _cancellation):
            if project_id == "P1": raise RuntimeError("unreadable")
            return good
        def load_current_page(self, page_id, _cancellation): return _document(page_id, 1, period.start, _body(page_id, "x"))
        def load_page_versions(self, page_id, _period, current, _cancellation): return (current,)

    batch = ConfluenceWeeklyAuditUseCase(Source()).run((failed, good), period)

    assert batch.projects[0].findings[0].status is AuditStatus.FAILED
    assert len(batch.projects[1].findings) == 8


def test_confluence_reuses_page_material_for_points_on_the_same_page() -> None:
    project = _project()
    period = AuditPeriod(datetime(2026, 8, 17, tzinfo=TZ), datetime(2026, 8, 24, tzinfo=TZ))
    current_calls, version_calls = [], []

    class Source:
        def load_project_details(self, _project_id, _details, _cancellation): return project
        def load_current_page(self, page_id, _cancellation):
            current_calls.append(page_id)
            return _document(page_id, 2, datetime(2026, 8, 20, tzinfo=TZ), _body(page_id, "new"))
        def load_page_versions(self, page_id, _period, current, _cancellation):
            version_calls.append(page_id)
            return (_document(page_id, 1, datetime(2026, 8, 10, tzinfo=TZ), _body(page_id, "old")), current)

    ConfluenceWeeklyAuditUseCase(Source()).run((project,), period)

    assert current_calls.count("10") == 1
    assert version_calls.count("10") == 1


def test_confluence_exporter_writes_one_workbook_per_product_line(tmp_path) -> None:
    period = AuditPeriod(datetime(2026, 8, 17, tzinfo=TZ), datetime(2026, 8, 24, tzinfo=TZ))
    projects = (_project("P1", "DOPL"), _project("P2", "TV"))

    class Source:
        def load_project_details(self, project_id, _details, _cancellation):
            return next(project for project in projects if project.identity.project_id == project_id)
        def load_current_page(self, page_id, _cancellation): return _document(page_id, 1, period.start, _body(page_id, "x"))
        def load_page_versions(self, page_id, _period, current, _cancellation): return (current,)

    batch = ConfluenceWeeklyAuditUseCase(Source()).run(projects, period)
    paths = export_audit_xlsx_by_product_line(batch, tmp_path)

    assert [path.name for path in paths] == [f"DOPL_{batch.id}.xlsx", f"TV_{batch.id}.xlsx"]
    assert load_workbook(paths[0]).sheetnames == ["Project Weekly Audit"]


def test_confluence_exporter_preserves_legacy_mode_sections_columns_sort_and_links(tmp_path) -> None:
    period = AuditPeriod(datetime(2026, 8, 17, tzinfo=TZ), datetime(2026, 8, 24, tzinfo=TZ))
    findings = tuple(
        AuditFinding("P", point.label, point.rule_id, AuditStatus.UPDATED, "")
        for point in UPDATE_MATRIX_POINTS
    )
    batch = AuditBatch("legacy-format", period, period.end, (
        ProjectAudit(_project("P2", support_mode="B", project_status="NORMAL", commercial_date="2026-03-01"), findings, ("Bob",)),
        ProjectAudit(_project("P3", support_mode="A", project_status="NORMAL", commercial_date="2025-03-01"), findings, ("Alice",)),
        ProjectAudit(_project("P1", support_mode="A", project_status="NORMAL", commercial_date="2026-03-01"), findings, ("Alice",)),
    ))

    [path] = export_audit_xlsx_by_product_line(batch, tmp_path)

    workbook = load_workbook(path)
    sheet = workbook.active
    assert [sheet.cell(1, column).value for column in range(1, 7)] == [
        "Support Mode", "A", "Project Status", "NORMAL", "审查周期",
        "2026-08-17 - 2026-08-24",
    ]
    assert [sheet.cell(2, column).value for column in range(1, 6)] == [
        "Owner", "年份", "项目名", "项目链接", UPDATE_MATRIX_POINTS[0].label,
    ]
    assert [sheet.cell(row, 3).value for row in (3, 4)] == ["Project P3", "Project P1"]
    assert sheet.cell(3, 4).hyperlink.target == "https://c/P3"
    assert [sheet.cell(5, column).value for column in range(1, 5)] == [
        "Support Mode", "B", "Project Status", "NORMAL",
    ]
    assert sheet.cell(6, 1).value == "Owner"
    assert sheet.cell(7, 3).value == "Project P2"
    workbook.close()


def _body(page_id, value):
    if page_id == "11":
        return f"<table><tr><td>Category</td></tr><tr><td>{value}</td></tr></table>"
    points = [point for point in UPDATE_MATRIX_POINTS if {
        "10": "test_information", "11": "test_plan", "12": "environment",
        "13": "experience", "14": "report_store",
    }.get(page_id) == point.source_page]
    return "".join(f"<h2>{point.standard_name}</h2><p>{value}</p>" for point in points) or f"<p>{value}</p>"


def test_confluence_exporter_preserves_specific_invalid_format_reasons(tmp_path) -> None:
    period = AuditPeriod(datetime(2026, 8, 17, tzinfo=TZ), datetime(2026, 8, 24, tzinfo=TZ))
    findings = (
        AuditFinding("P1", "Test Information", "test.weekly", AuditStatus.INVALID_FORMAT, "格式有误：查询不到Phase Status"),
        AuditFinding("P1", "Test Information", "test.summary", AuditStatus.INVALID_FORMAT, "MissingSection"),
    )
    batch = AuditBatch("reason-report", period, period.end, (ProjectAudit(_project(), findings),))

    paths = export_audit_xlsx_by_product_line(batch, tmp_path)

    workbook = load_workbook(paths[0])
    rows = list(workbook.active.values)
    row = next(row for row in rows if "Project P1" in row)
    assert row[4:6] == ("格式有误：查询不到Phase Status", "格式有误：MissingSection")
    workbook.close()


def test_test_plan_accepts_nonempty_test_module_column_as_category() -> None:
    category = next(point for point in UPDATE_MATRIX_POINTS if point.rule_id == "plan.test")

    region = extract_page_region(_document(
        "plan", 1, datetime(2026, 8, 24, tzinfo=TZ),
        "<table><tr><th rowspan='4'>测试<br/>模块</th><th>测试<br/>Owner</th></tr>"
        "<tr><td>week51&amp;52</td></tr><tr><td>2023/xx/xx</td></tr>"
        "<tr><td>系统功能</td><td>启动与升级</td></tr></table>",
    ), category)

    assert region.found
    assert "系统功能" in region.content


def test_test_plan_does_not_accept_empty_or_narrative_content_as_category() -> None:
    category = next(point for point in UPDATE_MATRIX_POINTS if point.rule_id == "plan.test")

    empty = extract_page_region(_document(
        "plan", 1, datetime(2026, 8, 24, tzinfo=TZ),
        "<table><tr><th>测试模块</th><th>测试内容</th></tr></table>",
    ), category)
    narrative = extract_page_region(_document(
        "plan", 1, datetime(2026, 8, 24, tzinfo=TZ),
        "<p>本项目按照里程碑完成测试与问题闭环。</p>",
    ), category)

    assert not empty.found
    assert not narrative.found


def test_confluence_use_case_starts_every_project_when_requests_are_below_capacity() -> None:
    from threading import Barrier
    from core.async_tasks import AsyncTaskManager

    projects = (_project("P1"), _project("P2"))
    period = AuditPeriod(datetime(2026, 8, 17, tzinfo=TZ), datetime(2026, 8, 24, tzinfo=TZ))
    barrier = Barrier(2, timeout=1)

    class Source:
        def load_project_details(self, project_id, _details, _cancellation):
            barrier.wait()
            return next(project for project in projects if project.identity.project_id == project_id)
        def load_current_page(self, page_id, _cancellation): return _document(page_id, 1, period.start, _body(page_id, "x"))
        def load_page_versions(self, _page_id, _period, current, _cancellation): return (current,)

    manager = AsyncTaskManager(max_workers=3)
    try:
        root = manager.submit_coordinator(
            "confluence-review", lambda token, _progress: ConfluenceWeeklyAuditUseCase(Source()).run(
                projects, period, task_manager=manager, parent_task_id=token.task_id,
            ),
        )
        batch = root.result(timeout=2)
        root_id = manager.task_id(root)
        assert [item.project.identity.project_id for item in batch.projects] == ["P1", "P2"]
        assert len([task for task in manager._tasks.values() if task.parent_id == root_id]) == 2
    finally:
        manager.close()


def test_confluence_root_cancellation_cancels_scheduled_project_children() -> None:
    from core.async_tasks import AsyncTaskManager, TaskCancelled

    projects = (_project("P1"), _project("P2"))
    period = AuditPeriod(datetime(2026, 8, 17, tzinfo=TZ), datetime(2026, 8, 24, tzinfo=TZ))
    manager = AsyncTaskManager(max_workers=3)
    root = manager.register_long_running("confluence-review")

    class Source:
        def load_project_details(self, project_id, _details, _cancellation):
            if project_id == "P1":
                manager.cancel(root)
            return next(project for project in projects if project.identity.project_id == project_id)
        def load_current_page(self, page_id, _cancellation):
            return _document(page_id, 1, period.start, _body(page_id, "x"))
        def load_page_versions(self, _page_id, _period, current, _cancellation): return (current,)

    try:
        with pytest.raises(TaskCancelled):
            ConfluenceWeeklyAuditUseCase(Source()).run(
                projects, period, task_manager=manager, parent_task_id=root,
            )
        assert all(task.token._event.is_set() for task in manager._tasks.values() if task.root_id == root)
    finally:
        manager.close()


def test_confluence_review_uses_all_three_available_detail_workers() -> None:
    from threading import Barrier
    from core.async_tasks import AsyncTaskManager

    projects = (_project("P1"), _project("P2"), _project("P3"))
    period = AuditPeriod(datetime(2026, 8, 17, tzinfo=TZ), datetime(2026, 8, 24, tzinfo=TZ))
    barrier = Barrier(3, timeout=1)

    class Source:
        def load_project_details(self, project_id, _details, _cancellation):
            barrier.wait()
            return next(project for project in projects if project.identity.project_id == project_id)
        def load_current_page(self, page_id, _cancellation): return _document(page_id, 1, period.start, _body(page_id, "x"))
        def load_page_versions(self, _page_id, _period, current, _cancellation): return (current,)

    manager = AsyncTaskManager(max_workers=3)
    try:
        root = manager.submit_coordinator(
            "confluence-review", lambda token, _progress: ConfluenceWeeklyAuditUseCase(Source()).run(
                projects, period, task_manager=manager, parent_task_id=token.task_id,
            ),
        )
        batch = root.result(timeout=2)
        assert all(finding.status is not AuditStatus.FAILED for audit in batch.projects for finding in audit.findings)
    finally:
        manager.close()
