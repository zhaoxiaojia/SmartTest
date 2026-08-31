from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.reporting.excel import clean_excel_value, write_excel_workbook

from .models import AuditReport
from .rules import creator_names


_SUMMARY_WIDTHS = (28, 24, 85)
_DETAIL_WIDTHS = (14, 38, 10, 34, 18, 55, 22, 65, 38, 38)
_DETAIL_HEADERS = (
    "Issue Key",
    "Issue URL",
    "检查结果",
    "规则编号",
    "规范章节",
    "规范要求/标准格式",
    "Jira 字段",
    "当前 Jira 内容",
    "失败原因",
    "修改建议",
)
_STATIC_FAILURE_REASONS = {
    "COMPONENT.REQUIRED": "Component 为空。",
    "DESCRIPTION.STEPS_TO_REPRODUCE": "复现步骤缺失或为空。",
    "DESCRIPTION.ACTUAL_RESULTS": "实际结果缺失或为空。",
    "DESCRIPTION.EXPECTED_RESULTS": "预期结果缺失或为空。",
    "DESCRIPTION.COMPARISON": "版本对比信息缺失或为空。",
    "DESCRIPTION.NOTES": "备注缺失或为空。",
    "DESCRIPTION.RATE_FORMAT": "Description 中的复现概率格式无效。",
    "DESCRIPTION.NOTES_HW": "Notes 缺少硬件信息。",
    "DESCRIPTION.NOTES_SW": "Notes 缺少软件信息。",
}
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
_BODY_FONT = Font(name="Calibri", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="top", wrap_text=True)
_BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def export_audit_xlsx(
    report: AuditReport, *, output_path: Path,
    now: datetime | None = None,
) -> Path:
    del now
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    summary_rows, summary_headers = _summary_rows(report)
    detail_rows, detail_merges = _detail_rows(report)
    def populate(workbook):
        summary = workbook.active
        summary.title = "汇总"
        detail = workbook.create_sheet("违规明细")
        _populate_sheet(
            summary,
            summary_rows,
            widths=_SUMMARY_WIDTHS,
            header_rows=summary_headers,
        )
        _populate_sheet(
            detail,
            detail_rows,
            widths=_DETAIL_WIDTHS,
            header_rows={1},
            auto_filter=True,
            merged_ranges=detail_merges,
        )

    return write_excel_workbook(target, populate)


def _summary_rows(report: AuditReport) -> tuple[list[list[object]], set[int]]:
    qa_creator_names = creator_names()
    failed_by_creator: dict[str, set[str]] = {}
    for issue in report.issues:
        if not issue.passed:
            creator = str(issue.creator or "未知创建人")
            failed_by_creator.setdefault(creator, set()).add(issue.key)
    passed = report.passed_count
    rows: list[list[object]] = [
        ["指标", "值"],
        ["生成时间", report.generated_at.isoformat()],
        ["JQL 查询条件", report.resolved.jql],
        ["问题总数", report.total_count],
        ["通过 Jira 数", passed],
        ["不通过 Jira 数", report.total_count - passed],
        [
            "通过率",
            f"{(passed / report.total_count * 100) if report.total_count else 0:.2f}%",
        ],
        [],
        ["创建人", "违规 Jira 数量", "违规 Jira 号"],
    ]
    rows.extend(
        [creator, len(keys), "、".join(sorted(keys))]
        for creator, keys in sorted(failed_by_creator.items())
    )
    rows.append([])
    whitelist_title_row = len(rows) + 1
    rows.append(
        [
            "审查 Creator 白名单",
            f"共 {len(qa_creator_names)} 人",
            "工具仅抓取 Creator 命中以下名单的 Jira",
        ]
    )
    whitelist_header_row = len(rows) + 1
    rows.append(["序号", "Creator 姓名", ""])
    rows.extend(
        [index, creator_name, ""]
        for index, creator_name in enumerate(sorted(qa_creator_names), 1)
    )
    return rows, {1, 9, whitelist_title_row, whitelist_header_row}


def _detail_rows(report: AuditReport) -> tuple[list[list[object]], tuple[str, ...]]:
    rules_by_id = {rule.rule_id: rule for rule in report.rules}
    rows: list[list[object]] = [list(_DETAIL_HEADERS)]
    merged_ranges: list[str] = []
    for issue in report.issues:
        first_row = len(rows) + 1
        for index, violation in enumerate(issue.violations):
            rule = rules_by_id.get(violation.rule_id)
            requirement = rule.requirement if rule else ""
            guidance = (
                violation.guidance
                if violation.rule_id
                in {"SUMMARY.FORMAT", "SUMMARY.PROBABILITY"}
                else rule.guidance if rule else violation.guidance
            )
            rows.append(
                [
                    issue.key if index == 0 else "",
                    issue.url if index == 0 else "",
                    "不符合" if index == 0 else "",
                    violation.rule_id,
                    violation.section,
                    requirement,
                    violation.field,
                    violation.observed,
                    _STATIC_FAILURE_REASONS.get(
                        violation.rule_id, violation.reason
                    ),
                    guidance,
                ]
            )
        if len(issue.violations) > 1:
            last_row = first_row + len(issue.violations) - 1
            merged_ranges.extend(
                f"{column}{first_row}:{column}{last_row}"
                for column in ("A", "B", "C")
            )
    return rows, tuple(merged_ranges)


def _populate_sheet(
    sheet,
    rows: list[list[object]],
    *,
    widths: tuple[int, ...],
    header_rows: set[int],
    auto_filter: bool = False,
    merged_ranges: tuple[str, ...] = (),
) -> None:
    for row_number, row in enumerate(rows, 1):
        cleaned = [clean_excel_value(value) for value in row]
        sheet.append(cleaned)
        content_length = max((len(str(value)) for value in cleaned), default=0)
        sheet.row_dimensions[row_number].height = (
            30 if row_number in header_rows
            else min(90, max(18, 15 * max(1, content_length // 80 + 1)))
        )
        for column, value in enumerate(cleaned, 1):
            cell = sheet.cell(row_number, column)
            if isinstance(value, str):
                cell.data_type = "s"
            if row_number in header_rows:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGNMENT
            else:
                cell.font = _BODY_FONT
                cell.alignment = _BODY_ALIGNMENT
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    if auto_filter:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{len(rows)}"
    for merged_range in merged_ranges:
        sheet.merge_cells(merged_range)
