from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .workbook import clean_excel_value, write_excel_workbook


_HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
_STATUS_FILLS = {
    "已更新": PatternFill("solid", fgColor="FFC6EFCE"),
    "未更新": PatternFill("solid", fgColor="FFFFEB9C"),
    "格式有误": PatternFill("solid", fgColor="FFFFC7CE"),
}


def _status_fill(value):
    fill = _STATUS_FILLS.get(value)
    if fill is None and isinstance(value, str) and value.startswith("格式有误："):
        fill = _STATUS_FILLS["格式有误"]
    return fill


def write_xlsx_table(
    output_path,
    *,
    sheet_name,
    headers,
    rows,
    hyperlinks=None,
):
    """Write one reusable, styled XLSX table through the global report owner."""
    def populate(workbook):
        _populate_table(
            workbook,
            sheet_name=sheet_name,
            headers=headers,
            rows=rows,
            hyperlinks=hyperlinks,
        )

    return write_excel_workbook(output_path, populate)


def _populate_table(workbook, *, sheet_name, headers, rows, hyperlinks):
    sheet = workbook.active
    sheet.title = str(sheet_name)
    sheet.append([clean_excel_value(value) for value in headers])
    for row in rows:
        sheet.append([clean_excel_value(value) for value in row])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            fill = _status_fill(cell.value)
            if fill is not None:
                cell.fill = fill
    for (row, column), url in (hyperlinks or {}).items():
        cell = sheet.cell(row=row, column=column)
        cell.hyperlink = str(url)
        cell.style = "Hyperlink"
    for index, column in enumerate(sheet.columns, 1):
        width = max(
            (len(str(cell.value or "")) for cell in column),
            default=0,
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(
            45, max(12, width + 2),
        )



def write_xlsx_sections(
    output_path, *, sheet_name, sections, wrap_data=True,
):
    """Write repeated group/header/data sections with one global XLSX owner."""
    def populate(workbook):
        _populate_sections(
            workbook,
            sheet_name=sheet_name,
            sections=sections,
            wrap_data=wrap_data,
        )

    return write_excel_workbook(output_path, populate)


def _populate_sections(workbook, *, sheet_name, sections, wrap_data=True):
    sheet = workbook.active
    sheet.title = str(sheet_name)
    hyperlinks = {}
    header_rows = []
    group_rows = []
    for section in sections:
        sheet.append([clean_excel_value(value) for value in section["group"]])
        group_rows.append(sheet.max_row)
        sheet.append([clean_excel_value(value) for value in section["headers"]])
        header_rows.append(sheet.max_row)
        for row_index, row in enumerate(section["rows"]):
            sheet.append([clean_excel_value(value) for value in row])
            for column, url in section.get("hyperlinks", {}).get(row_index, {}).items():
                hyperlinks[(sheet.max_row, column)] = url

    sheet.freeze_panes = "A3"
    for row_index in group_rows:
        for cell in sheet[row_index]:
            cell.fill = PatternFill("solid", fgColor="FFD9EAF7")
            cell.font = Font(name="Calibri", size=11, bold=True, color="FF1F1F1F")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_index in header_rows:
        for cell in sheet[row_index]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
            )
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row not in header_rows and cell.row not in group_rows:
                cell.alignment = Alignment(
                    vertical="top", wrap_text=wrap_data,
                )
                fill = _status_fill(cell.value)
                if fill is not None:
                    cell.fill = fill
    for (row, column), url in hyperlinks.items():
        cell = sheet.cell(row=row, column=column)
        cell.hyperlink = str(url)
        cell.style = "Hyperlink"
    for index, column in enumerate(sheet.columns, 1):
        width = max((len(str(cell.value or "")) for cell in column), default=0)
        sheet.column_dimensions[get_column_letter(index)].width = min(
            55, max(12, width + 2),
        )
