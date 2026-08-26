from __future__ import annotations

import hashlib
import os
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from core.config.jsonTool import app_data_dir


SOURCES = {"jira", "confluence"}


class ReportNotFoundError(LookupError):
    pass


def _default_jira_dir() -> Path:
    return Path.home() / "Downloads"


def _default_confluence_dir() -> Path:
    return app_data_dir() / "confluence_audit" / "reports"


class ClientAuditReportOwner:
    """Read-only adapter for XLSX files exported by the Client audit owners."""

    def __init__(self, *, jira_dir: Path | None, confluence_dir: Path | None):
        self._roots = {
            "jira": Path(jira_dir).resolve() if jira_dir is not None else None,
            "confluence": Path(confluence_dir).resolve() if confluence_dir is not None else None,
        }

    @classmethod
    def from_environment(cls):
        jira = os.getenv("SMARTTEST_JIRA_REPORT_DIR")
        confluence = os.getenv("SMARTTEST_CONFLUENCE_REPORT_DIR")
        return cls(
            jira_dir=Path(jira) if jira else _default_jira_dir(),
            confluence_dir=Path(confluence) if confluence else _default_confluence_dir(),
        )

    def list_reports(self, source: str, filters: dict | None = None) -> dict:
        root = self._root(source)
        if root is None:
            return self._response("config_missing", [], 0)
        if not root.exists():
            return self._response("empty", [], 0)
        reports, failures = [], 0
        for path in sorted(root.glob(self._pattern(source)), key=lambda item: item.stat().st_mtime, reverse=True):
            try:
                detail = self._read(source, path)
                reports.append({key: detail[key] for key in (
                    "id", "title", "reportType", "productLine", "project", "year",
                    "generatedAt", "status", "sourceUrl", "downloadName", "jql",
                )})
            except Exception:
                failures += 1
        facets = {
            "productLines": sorted({row["productLine"] for row in reports if row["productLine"]}),
            "years": sorted({row["year"] for row in reports if row["year"]}),
            "reportTypes": sorted({row["reportType"] for row in reports if row["reportType"]}),
        }
        visible = self._filter(reports, filters or {})
        state = "partial_success" if failures else "ready" if visible else "empty"
        return self._response(state, visible, failures, facets)

    def get_report(self, source: str, report_id: str) -> dict:
        path = self._find(source, report_id)
        return self._read(source, path)

    def download_path(self, source: str, report_id: str) -> Path:
        return self._find(source, report_id)

    def _root(self, source: str) -> Path | None:
        if source not in SOURCES:
            raise ReportNotFoundError(source)
        return self._roots[source]

    @staticmethod
    def _pattern(source: str) -> str:
        return "jira_format_audit_*.xlsx" if source == "jira" else "*.xlsx"

    def _find(self, source: str, report_id: str) -> Path:
        root = self._root(source)
        if root is None or not root.exists():
            raise ReportNotFoundError(report_id)
        for path in root.glob(self._pattern(source)):
            if self._id(path) == report_id:
                return path.resolve()
        raise ReportNotFoundError(report_id)

    def _read(self, source: str, path: Path) -> dict:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sections = []
        try:
            for sheet in workbook.worksheets:
                rows = [[self._cell(value) for value in row] for row in sheet.iter_rows(values_only=True)]
                rows = [row for row in rows if any(value not in (None, "") for value in row)]
                headers = rows[0] if rows else []
                sections.append({"title": sheet.title, "headers": headers, "rows": rows[1:]})
        finally:
            workbook.close()
        summary = self._summary(source, sections)
        metadata = self._metadata(source, path, sections)
        metadata["status"] = "attention" if summary["attention"] or summary["failed"] else "completed"
        return {**metadata, "summary": summary, "sections": sections}

    def _metadata(self, source: str, path: Path, sections: list[dict]) -> dict:
        timestamp = datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat()
        source_url = self._find_url(sections)
        jql = self._named_value(sections, "JQL 查询条件") if source == "jira" else ""
        if source == "jira":
            title, report_type, product_line = "Jira format audit", "Issue Audit", ""
        else:
            product_line = path.stem.split("_", 1)[0]
            title, report_type = f"{product_line} project weekly audit", "Weekly Audit"
        projects = []
        for section in sections:
            headers = [str(value or "").strip() for value in section["headers"]]
            if "项目名" in headers:
                column = headers.index("项目名")
                projects.extend(str(row[column]) for row in section["rows"] if len(row) > column and row[column])
        years = [int(value) for section in sections for row in section["rows"] for value in row if isinstance(value, (int, float)) and 2000 <= value <= 2200]
        return {
            "id": self._id(path), "title": title, "reportType": report_type,
            "productLine": product_line, "project": ", ".join(dict.fromkeys(projects)), "year": max(years, default=None),
            "generatedAt": timestamp, "status": "completed", "sourceUrl": source_url,
            "downloadName": path.name, "jql": str(jql or ""),
        }

    @staticmethod
    def _summary(source: str, sections: list[dict]) -> dict:
        values = {}
        for section in sections:
            for row in section["rows"]:
                if len(row) >= 2:
                    values[str(row[0]).strip()] = row[1]
        if source == "jira":
            total = int(values.get("问题总数") or 0)
            passed = int(values.get("通过 Jira 数") or 0)
            failed = int(values.get("不通过 Jira 数") or max(total - passed, 0))
            return {"total": total, "passed": passed, "attention": failed, "failed": 0}
        return {"total": None, "passed": None, "attention": None, "failed": None}

    @staticmethod
    def _find_url(sections: list[dict]) -> str:
        for section in sections:
            for row in section["rows"]:
                for value in row:
                    if isinstance(value, str) and re.match(r"^https?://", value):
                        return value
        return ""

    @staticmethod
    def _filter(reports: list[dict], filters: dict) -> list[dict]:
        product_line = str(filters.get("product_line") or "").casefold()
        report_type = str(filters.get("report_type") or "").casefold()
        search = str(filters.get("search") or "").casefold()
        year = filters.get("year")
        jql = " ".join(str(filters.get("jql") or "").split()).casefold()
        return [row for row in reports if
                (not product_line or row["productLine"].casefold() == product_line) and
                (not report_type or row["reportType"].casefold() == report_type) and
                (not year or row["year"] == int(year)) and
                (not jql or " ".join(row["jql"].split()).casefold() == jql) and
                (not search or search in " ".join(str(value) for value in row.values()).casefold())]

    @staticmethod
    def _named_value(sections: list[dict], name: str):
        for section in sections:
            for row in section["rows"]:
                if len(row) >= 2 and str(row[0] or "").strip() == name:
                    return row[1]
        return None

    @staticmethod
    def _response(state, reports, failures, facets=None):
        return {"state": state, "reports": reports, "failures": failures,
                "facets": facets or {"productLines": [], "years": [], "reportTypes": []}}

    @staticmethod
    def _id(path: Path) -> str:
        return hashlib.sha256(str(path.resolve()).encode("utf-8")).hexdigest()[:20]

    @staticmethod
    def _cell(value):
        if isinstance(value, datetime):
            return value.isoformat()
        return value
